import time
from datetime import datetime

import fuzzywuzzy.utils
import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

excel_sheet = "healthcare_health.xlsx"
rooms = ["General ward", "Semi private", "Single private room", "deluxe room"]
rooms_dict = {}
st = datetime.now()
master_df: pd.DataFrame = pd.read_excel("BIlling Code Master.xlsx", sheet_name="IRDA Updated", converters={'SUB_CATEGORY_CODE':str,'CATEGORY_CODE':str})
master_df = master_df.fillna("")
master_df = master_df.replace(to_replace="NaN", value="")
print((datetime.now() - st).total_seconds())
print(master_df.keys())
temp_room_df = master_df[master_df["SUB_CATEGORY_CODE"].str.contains("101000", na=False)]#roomcharges
print(temp_room_df)
names_list = temp_room_df["NAME"].str.lower().tolist()
print(names_list)
for room in rooms:
    print(room)
    matches = process.extractOne(room.lower(), names_list,
                              scorer=fuzz.partial_ratio)
    print(matches)
    index = names_list.index(matches[0])
    print(index)
    rooms_dict[room] = temp_room_df["ITEM_CODE"][index + 1]
    # for i in range(len(matches)):
    #
    #     if matches[i][1] > 85:
    #         print(matches[i])
    #         break
print(rooms_dict)

inter_tariff_df = pd.DataFrame()
for sheet in pd.ExcelFile("healthcare_health.xlsx").sheet_names:
    inter_tariff_df = inter_tariff_df.append(pd.read_excel("healthcare_health.xlsx", sheet_name=sheet))
print(inter_tariff_df)
inter_tariff_df = inter_tariff_df.fillna("")
inter_tariff_df = inter_tariff_df.replace("NA", "0.0")
services_list = inter_tariff_df["Name of the procedure/treatment"].tolist()
packages_items_df:pd.DataFrame = master_df
packages_list = packages_items_df["NAME"].str.lower().tolist()
package_indexes = packages_items_df.index.tolist()
# for i in range(len(services_list)):
#     print(inter_tariff_df["Name of the procedure/treatment"][i])
# print(packages_list)
# print(packages_items_df)
mappings_df = pd.read_excel("mappings.xlsx", sheet_name="Sheet1", converters={'Name': str})
mappings_list = mappings_df["Name"].str.lower().tolist()
print(type(mappings_list))
wb = load_workbook(r"template.xlsx")
ws = wb["Sheet1"]
j = 2

yellow_fill = PatternFill(start_color="feff00",
                   fill_type='solid')
for index, row in inter_tariff_df.iterrows():
    if not bool(str(row["Name of the procedure/treatment"]).lower().replace("\r", " ").replace("\n", " ")) or not bool(row["General ward"]):
        continue
    possiblities = []

    procedure_name = str(row["Name of the procedure/treatment"]).lower().replace("\r", " ").replace("\n", " ")
    i, z = process.extractOne(procedure_name, mappings_list, scorer=fuzz.ratio)
    if z > 95:
        procedure_name = mappings_df["Possibility"][mappings_list.index(i)]
    print("actual", str(row["Name of the procedure/treatment"]).lower().replace("\r", " ").replace("\n", " "))

    procedure_name = procedure_name.replace("therapeutic", "diagnostic/therapeutic").replace("charges", "charge")
    print("procedure", procedure_name)
    matches = process.extract(procedure_name, packages_list, scorer=fuzz.partial_token_sort_ratio, limit=30)
    print("first", matches)
    first_match = [match[0] for match in matches]
    inter_matches = process.extract(procedure_name, first_match, scorer=fuzz.ratio, limit=15)

    print(inter_matches)
    temp_list = procedure_name.lower().replace("(", "").replace(")", "").strip().split(" ")
    if "per" in temp_list:
        temp_list.remove("per")
    if "check" in temp_list:
        temp_list.remove("check")
    print("templist", temp_list)
    temp_match_list = []
    for match, percent in inter_matches:
        if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
            temp_match_list.append(match)
    if len(temp_match_list) == 0:
        temp_match_list = [match[0] for match in inter_matches]
    print(temp_match_list)
    matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.partial_ratio, limit=4)

    print(matches)
    # if matches[0][1] <= 70:
    #     matches = process.extract(procedure_name, [match[0] for match in inter_matches], scorer=fuzz.partial_ratio, limit=4)
    temp_str = ""
    temp_percent = 0
    for match, percent in matches:
        possiblities.append(match)
        if percent > temp_percent:
            temp_str = match
            # if len(match)/len(str(row["Name of the procedure/treatment"]).lower().replace("\n", "")) < 40.0:
            #     continue
            if percent == 100:
                temp_percent = percent
                break

            temp_percent = percent
    if temp_str == "":
        temp_str = matches[0][0]
    if temp_percent > 85:
        possiblities = []
    print(temp_str)
    master_index = packages_list.index(temp_str)
    master_index = int(package_indexes[master_index])
    print(master_index)
    print(packages_items_df["ITEM_CODE"][4265])
    for room in rooms:
        ws.cell(row=j, column=1).value = "Kanpur Health care"
        ws.cell(row=j, column=2).value = ""
        ws.cell(row=j, column=3).value = row["Name of the procedure/treatment"]
        ws.cell(row=j, column=4).value = packages_items_df["ITEM_CODE"][master_index]
        ws.cell(row=j, column=5).value = packages_items_df["NAME"][master_index]
        ws.cell(row=j, column=6).value = ""
        ws.cell(row=j, column=7).value = ""
        ws.cell(row=j, column=8).value = rooms_dict[room]
        ws.cell(row=j, column=9).value = row[room]
        ws.cell(row=j, column=10).value = ",".join(possiblities[1:])
        if bool(",".join(possiblities[1:])):
            ws.cell(row=j, column=10).fill = yellow_fill
        print(possiblities)
        j += 1

wb.save(r"kanpur_output.xlsx")
wb.close()