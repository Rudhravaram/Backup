import time
from datetime import date
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from Main_icd_code_generation import main_icd10, main_icd10_pcs


def clean_list(list_: list):
    temp_list = []
    for ele in list_:
        if bool(str(ele).strip()):
            temp_list.append(str(ele).strip())
    return temp_list

def get_thevalues():
    master_file_name = r'D:\Tariff Matching Master\dev.xlsx'
    rooms_dict = {}
    st = datetime.now()
    master_df: pd.DataFrame = pd.read_excel(master_file_name, sheet_name="Sheet4",
                                            converters={'SUB_CATEGORY_CODE': str, 'CATEGORY_CODE': str})
    master_df = master_df.fillna("")
    master_df = master_df.replace(to_replace="NaN", value="")
    print((datetime.now() - st).total_seconds())
    print(master_df.keys())
    packages_items_df: pd.DataFrame = master_df
    packages_list = packages_items_df["NAME"].str.lower().tolist()
    package_indexes = packages_items_df.index.tolist()
    wb = load_workbook(master_file_name)
    ws = wb["Sheet4"]
    KM = 2
    today = date.today()
    print("Today's date:", today)
    print(datetime.now())
    for Pack in packages_list:
        Result=[]
        Z_results=[]
        Temp_pack=Pack
        Pack=str(Pack).lower().replace('pcp','Pneumocystis carinii pneumonia(PCP)').replace('charges','')
        try:
            print('+++++++++++++++++++++++++start++++++++++++++++++++++++++++++')
            print("Pack", Pack)
            main_words,code_list,desc,block,block_desc=main_icd10(Pack)
            code_list = clean_list(code_list)
            print('main_words',main_words)
            print('code_list',code_list)
            if code_list:
                for i in range(len(code_list)):
                    if not str(code_list[i]).lower().startswith('w') and not str(code_list[i]).lower().startswith('x') and not str(code_list[i]).lower().startswith('y') and not str(code_list[i]).lower().startswith('z'):
                        Result.append(code_list[i])
                        Z_results = []
                    else:
                        try:
                            final_code, final_input = main_icd10_pcs(main_words[i])
                            for j in final_code:
                                Z_results = []
                                Result.append(j)
                        except:
                            try:
                                if len(code_list)==1:
                                    print('Len=1')
                                    Result.append(code_list[0])
                                else:
                                    Z_results.append(code_list[i])
                            except:
                                print('None')
            else:
                try:
                    final_code, final_input = main_icd10_pcs(Pack)
                    for ji in final_code:
                        Z_results = []
                        Result.append(ji)
                except:
                    print(Pack)
        except:
            print('Sleep')
            time.sleep(5)
            main_words, code_list, desc, block, block_desc = main_icd10(Pack)
            code_list = clean_list(code_list)
            print('main_words', main_words)
            print('code_list', code_list)
            if code_list:
                for i in range(len(code_list)):
                    if not str(code_list[i]).lower().startswith('w') and not str(code_list[i]).lower().startswith(
                            'x') and not str(code_list[i]).lower().startswith('y') and not str(
                            code_list[i]).lower().startswith('z'):
                        Result.append(code_list[i])
                        Z_results = []
                    else:
                        try:
                            final_code, final_input = main_icd10_pcs(main_words[i])
                            for j in final_code:
                                Z_results = []
                                Result.append(j)
                        except:
                            try:
                                if len(code_list) == 1:
                                    print('Len=1')
                                    Result.append(code_list[0])
                                else:
                                    Z_results.append(code_list[i])
                            except:
                                print('None')
            else:
                try:
                    final_code, final_input = main_icd10_pcs(Pack)
                    for ji in final_code:
                        Z_results = []
                        Result.append(ji)
                except:
                    print(Pack)
        print('Result', Result)
        master_index = packages_list.index(Temp_pack)
        master_index = int(package_indexes[master_index])
        id_codes = ''
        Result=clean_list(Result)
        if not Result:
            if Z_results:
                for i in Z_results:
                    Result.append(i)

        if Result:
            for i in Result:
                tem_t = False
                if not id_codes:
                    tem_t = True
                    id_codes = str(i)
                if id_codes and not tem_t:
                    id_codes = id_codes + ', ' + str(i)
        print("id_codes",id_codes)
        print('+++++++++++++++++++++++++End++++++++++++++++++++++++++++++')
        ws.cell(row=KM, column=5).value = id_codes
        KM += 1
    print("Today's date:", today)
    print(datetime.now())
    wb.save(r"dev.xlsx")
    wb.close()
get_thevalues()