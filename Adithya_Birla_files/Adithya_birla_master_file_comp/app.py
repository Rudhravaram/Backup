import time
from datetime import datetime
import fuzzywuzzy.utils
import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from googlesearch import search




def get_data():
    excel_sheet = r"D:\Tariff Matching Master\HS00017509-Rate List\List.xlsx"
    master_file_name = r'D:\Tariff Matching Master\HS00017509-Rate List\BIlling Code Master.xlsx'
    mappings_fileName = r'D:\Tariff Matching Master\Reference\mappings.xlsx'
    rooms = ["General ward", "SEMI PRIVATE", "PRIVATE", "DELUXE"]
    rooms_dict = {}
    st = datetime.now()

    master_df: pd.DataFrame = pd.read_excel(master_file_name, sheet_name="IRDA Updated",
                                            converters={'SUB_CATEGORY_CODE': str, 'CATEGORY_CODE': str})

    master_df = master_df.fillna("")
    master_df = master_df.replace(to_replace="NaN", value="")
    print((datetime.now() - st).total_seconds())
    print(master_df.keys())

    temp_room_df = master_df[master_df["SUB_CATEGORY_CODE"].str.contains("101000", na=False)]  # roomcharges
    print("temp_room_df", temp_room_df)
    names_list = temp_room_df["NAME"].str.lower().tolist()
    print("names_list", names_list)
    for room in rooms:
        print(room)
        matches = process.extractOne(room.lower(), names_list,
                                     scorer=fuzz.partial_ratio)
        print(matches)
        index = names_list.index(matches[0])
        print(index)
        rooms_dict[room] = temp_room_df["ITEM_CODE"][index + 1]

    print("rooms_dict", rooms_dict)

    inter_tariff_df = pd.DataFrame()
    for sheet in pd.ExcelFile(excel_sheet).sheet_names:
        inter_tariff_df = inter_tariff_df.append(pd.read_excel(excel_sheet, sheet_name=sheet))
    inter_tariff_df = inter_tariff_df.fillna("")
    inter_tariff_df = inter_tariff_df.replace("NA", "0.0")
    inter_tariff_df.to_excel("temp.xlsx")
    services_list = inter_tariff_df["DESCRIPTION"].tolist()
    print("services_list", services_list)
    packages_items_df: pd.DataFrame = master_df
    packages_list = packages_items_df["NAME"].str.lower().tolist()
    ICD_CODEss = packages_items_df["ICD_CODE"].str.lower().tolist()
    package_indexes = packages_items_df.index.tolist()
    print("packages_list", packages_list)
    print("package_indexes", package_indexes)

    mappings_df = pd.read_excel(mappings_fileName, sheet_name="Sheet1", converters={'Name': str})
    mappings_list = mappings_df["Name"].str.lower().tolist()

    print(type(mappings_list))

    wb = load_workbook(r"D:\Tariff Matching Master\HS00017509-Rate List\template.xlsx")
    ws = wb["Sheet1"]
    j = 2
    yellow_fill = PatternFill(start_color="feff00",
                              fill_type='solid')

    for index, row in inter_tariff_df.iterrows():
        if not bool(str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")):
            continue
        possiblities = []

        procedure_name = str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")
        print(type(procedure_name))
        i, z = process.extractOne(procedure_name, mappings_list, scorer=fuzz.ratio)
        print("i", i)
        print("z", z)

        if z > 95:
            procedure_name = mappings_df["Possibility"][mappings_list.index(i)]
        print("actual", str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " "))

        procedure_name = procedure_name.replace("therapeutic", "diagnostic/therapeutic").replace("charges", "charge")
        print("procedure", procedure_name)
        id_code = icd10_code(procedure_name)
        matches = process.extract(procedure_name, packages_list, scorer=fuzz.partial_token_sort_ratio, limit=30)
        print("first", matches)
        first_match = [match[0] for match in matches]
        print('first_match', first_match)
        inter_matches = process.extract(procedure_name, first_match, scorer=fuzz.partial_ratio, limit=15)
        print("inter_matches", inter_matches)
        temp_match_list1 = []
        temp_list1 = id_code
        vICD_lis=[]
        for match, percent in inter_matches:
            master_index = packages_list.index(match)
            master_index = int(package_indexes[master_index])
            IC_DCODE=packages_items_df["ICD_CODE"][master_index]
            if IC_DCODE in temp_list1:
                temp_match_list1.append(match)
                vICD_lis.append(IC_DCODE)
        print("id_code",id_code)
        print("temp_match_list1",temp_match_list1)
        print("vICD_lis",vICD_lis)
        if temp_match_list1:
            temp_percent = 0
            temp_str = ""
            matches = process.extract(procedure_name, temp_match_list1, scorer=fuzz.ratio, limit=4)
            print("matches",matches)
            for match, percent in matches:
                possiblities.append(match)
                if percent > temp_percent:
                    temp_str = match
                    if percent == 100:
                        temp_percent = percent
                        break
                    temp_percent = percent
            if temp_str == "":
                temp_str = matches[0][0]
            if temp_percent > 85:
                possiblities = []
            print(temp_str)
            master_index = packages_list.index(temp_str)
            master_index = int(package_indexes[master_index])
            print("master_index",master_index)
        else:
            temp_list = procedure_name.lower().replace("(", "").replace(")", "").strip().split(" ")
            if "per" in temp_list:
                temp_list.remove("per")
            if "check" in temp_list:
                temp_list.remove("check")
            print("templist", temp_list)
            temp_match_list = []
            for match, percent in inter_matches:
                if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                    temp_match_list.append(match)
            if len(temp_match_list) == 0:
                temp_match_list = [match[0] for match in inter_matches]
            print("temp_match_list",temp_match_list)
            matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.ratio, limit=10)
            print('matches',matches)
            temp_match_list = []
            for match, percent in matches:
                if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                    temp_match_list.append(match)
            if len(temp_match_list) == 0:
                temp_match_list = [match[0] for match in matches]

            matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.partial_ratio, limit=5)

            print("matchesaasaaa",matches)
            if matches[0][1] <= 70:
                matches = process.extract(procedure_name, [match[0] for match in matches], scorer=fuzz.partial_ratio, limit=4)
            temp_str = ""
            temp_percent = 0
            for match, percent in matches:
                possiblities.append(match)
                if percent > temp_percent:
                    temp_str = match
                    if percent == 100:
                        temp_percent = percent
                        break

                    temp_percent = percent
            if temp_str == "":
                temp_str = matches[0][0]
            if temp_percent > 85:
                possiblities = []
            print(temp_str)
            master_index = packages_list.index(temp_str)
            master_index = int(package_indexes[master_index])
            print(master_index)
        for room in rooms:
            ws.cell(row=j, column=1).value = "FARIDABAD MEDICAL CENTRE"
            ws.cell(row=j, column=2).value = ""
            ws.cell(row=j, column=3).value = row["DESCRIPTION"]
            ws.cell(row=j, column=4).value = id_code
            ws.cell(row=j, column=5).value = packages_items_df["ITEM_CODE"][master_index]
            ws.cell(row=j, column=6).value = packages_items_df["NAME"][master_index]
            ws.cell(row=j, column=7).value = ""
            ws.cell(row=j, column=8).value = ""
            ws.cell(row=j, column=9).value = rooms_dict[room]
            ws.cell(row=j, column=10).value = row[room.replace('General ward', 'ECONOMY')]
            ws.cell(row=j, column=11).value = ",".join(possiblities[1:])
            ws.cell(row=j, column=12).value = packages_items_df["ICD_CODE"][master_index]
            if bool(",".join(possiblities[1:])):
                ws.cell(row=j, column=11).fill = yellow_fill
            print(possiblities)
            j += 1
    wb.save(r"CENTRE_OUTPUT.xlsx")
    wb.close()

def icd10_code(dieases_data):
    # dieases_data = input("diagnosis name : ")
    query = str(dieases_data) + " icd 10"
    list_url = []
    for i in search(query, lang='en'):
        list_url.append(i)
    for i in range(len(list_url)):
        if str(list_url[i]).__contains__('icd10data'):
        # print(list_url[i])
            my_split = str(list_url[i]).split('/')[-1]
            return my_split



get_data()







