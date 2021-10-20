import time
from datetime import datetime
import fuzzywuzzy.utils
import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from googlesearch import search
from Main_icd_code_generation import main_icd10,clean_list,main_icd10_pcs




def get_data():
    excel_sheet = r"D:\Tariff Matching Master\Matching After Fezz Ratio\Book1.xlsx"
    master_file_name = r'D:\Tariff Matching Master\Matching After Fezz Ratio\BIlling Code Master.xlsx'
    mappings_fileName = r'D:\Tariff Matching Master\Matching After Fezz Ratio\mappings.xlsx'
    rooms = ["General ward", "SEMI PRIVATE", "PRIVATE", "DELUXE"]
    rooms_dict = {}
    st = datetime.now()

    master_df: pd.DataFrame = pd.read_excel(master_file_name, sheet_name="IRDA Updated",
                                            converters={'SUB_CATEGORY_CODE': str, 'CATEGORY_CODE': str})

    master_df = master_df.fillna("")
    master_df = master_df.replace(to_replace="NaN", value="")
    print((datetime.now() - st).total_seconds())
    print(master_df.keys())

    temp_room_df = master_df[master_df["SUB_CATEGORY_CODE"].str.contains("101000", na=False)]  # roomcharges
    print("temp_room_df", temp_room_df)
    names_list = temp_room_df["NAME"].str.lower().tolist()
    print("names_list", names_list)
    for room in rooms:
        print(room)
        matches = process.extractOne(room.lower(), names_list,
                                     scorer=fuzz.partial_ratio)
        print(matches)
        index = names_list.index(matches[0])
        print(index)
        rooms_dict[room] = temp_room_df["ITEM_CODE"][index + 1]

    print("rooms_dict", rooms_dict)

    inter_tariff_df = pd.DataFrame()
    for sheet in pd.ExcelFile(excel_sheet).sheet_names:
        inter_tariff_df = inter_tariff_df.append(pd.read_excel(excel_sheet, sheet_name=sheet))
    inter_tariff_df = inter_tariff_df.fillna("")
    inter_tariff_df = inter_tariff_df.replace("NA", "0.0")
    inter_tariff_df.to_excel("temp.xlsx")
    services_list = inter_tariff_df["DESCRIPTION"].tolist()
    print("services_list", services_list)
    packages_items_df: pd.DataFrame = master_df
    packages_list = packages_items_df["NAME"].str.lower().tolist()
    ICD_CODEss = packages_items_df["ICD_CODE"].str.lower().tolist()
    package_indexes = packages_items_df.index.tolist()
    print("packages_list", packages_list)
    print("package_indexes", package_indexes)

    mappings_df = pd.read_excel(mappings_fileName, sheet_name="Sheet1", converters={'Name': str})
    mappings_list = mappings_df["Name"].str.lower().tolist()

    print(type(mappings_list))

    wb = load_workbook(r"D:\Tariff Matching Master\Matching After Fezz Ratio\template.xlsx")
    ws = wb["Sheet1"]
    KM = 2
    yellow_fill = PatternFill(start_color="feff00",
                              fill_type='solid')

    for index, row in inter_tariff_df.iterrows():
        if not bool(str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")):
            continue
        possiblities = []

        procedure_name = str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")
        print(type(procedure_name))
        i, z = process.extractOne(procedure_name, mappings_list, scorer=fuzz.ratio)
        print("i", i)
        print("z", z)

        if z > 95:
            procedure_name = mappings_df["Possibility"][mappings_list.index(i)]
        print("actual", str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " "))

        procedure_name = procedure_name.replace("therapeutic", "diagnostic/therapeutic").replace("charges", "charge")
        print('------------Start-----------------------')
        print("procedure_name",procedure_name)
        main_words,code_list,desc,block,block_desc = main_icd10(procedure_name)
        code_list=clean_list(code_list)
        code_matche_list = []
        code_matches_Name = []
        id_codes = ''
        Master_id = ''
        Result = []
        Z_results=[]
        try:
            if code_list:
                for i in range(len(code_list)):
                    if not str(code_list[i]).lower().startswith('w') and not str(code_list[i]).lower().startswith(
                            'x') and not str(code_list[i]).lower().startswith('y') and not str(
                            code_list[i]).lower().startswith('z'):
                        Result.append(code_list[i])
                        Z_results = []
                    else:
                        try:
                            final_code, final_input = main_icd10_pcs(main_words[i])
                            for j in final_code:
                                Z_results = []
                                Result.append(j)
                        except:
                            try:
                                if len(code_list) == 1:
                                    print('Len=1')
                                    Result.append(code_list[0])
                                else:
                                    Z_results.append(code_list[i])
                            except:
                                print('None')
            else:
                try:
                    final_code, final_input = main_icd10_pcs(procedure_name)
                    final_code = clean_list(final_code)
                    if final_code:
                        for ji in final_code:
                            Z_results = []
                            Result.append(ji)
                except:
                    print("procedure_name",procedure_name)
        except:
            time.sleep(5)
            if code_list:
                for i in range(len(code_list)):
                    if not str(code_list[i]).lower().startswith('w') and not str(code_list[i]).lower().startswith(
                            'x') and not str(code_list[i]).lower().startswith('y') and not str(code_list[i]).lower().startswith('z'):
                        Result.append(code_list[i])
                        Z_results = []
                    else:
                        try:
                            final_code, final_input = main_icd10_pcs(main_words[i])
                            for j in final_code:
                                Z_results = []
                                Result.append(j)
                        except:
                            try:
                                if len(code_list) == 1:
                                    print('Len=1')
                                    Result.append(code_list[0])
                                else:
                                    Z_results.append(code_list[i])
                            except:
                                print('None')
            else:
                try:
                    final_code, final_input = main_icd10_pcs(procedure_name)
                    final_code = clean_list(final_code)
                    if final_code:
                        for ji in final_code:
                            Z_results = []
                            Result.append(ji)
                except:
                    print("procedure_name", procedure_name)
        Result = clean_list(Result)
        Z_results=clean_list(Z_results)
        if not Result:
            if Z_results:
                for i in Z_results:
                    Result.append(i)
        if Result:
            for i in Result:
                tem_t = False
                if not id_codes:
                    tem_t = True
                    id_codes = str(i)
                if id_codes and not tem_t:
                    id_codes = id_codes + ', ' + str(i)
            print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
            print("procedure", procedure_name)
            print("main_words", main_words)
            print("code_list", Result)
            print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
            matches = process.extract(procedure_name, packages_list, scorer=fuzz.partial_token_sort_ratio, limit=20)
            print("first", matches)
            first_match = [match[0] for match in matches]
            print('first_match', first_match)
            for match in first_match:
                Result_code_list_first_match=[]
                Z_Result_code_list_first_match=[]
                main_words_first_match, code_list_first_match, desc_first_match, block_first_match, block_desc_first_match = main_icd10(match)
                code_list_first_match=clean_list(code_list_first_match)
                try:
                    if code_list_first_match:
                        for i in range(len(code_list_first_match)):
                            if not str(code_list_first_match[i]).lower().startswith('w') and not str(code_list_first_match[i]).lower().startswith('x') and not str(code_list_first_match[i]).lower().startswith('y') and not str(code_list_first_match[i]).lower().startswith('z'):
                                Result_code_list_first_match.append(code_list_first_match[i])
                                Z_Result_code_list_first_match = []
                            else:
                                try:
                                    final_code, final_input = main_icd10_pcs(main_words_first_match[i])
                                    for j in final_code:
                                        Z_Result_code_list_first_match = []
                                        Result_code_list_first_match.append(j)
                                except:
                                    try:
                                        if len(code_list_first_match) == 1:
                                            print('Len=1')
                                            Result_code_list_first_match.append(code_list_first_match[0])
                                        else:
                                            Z_Result_code_list_first_match.append(code_list_first_match[i])
                                    except:
                                        print('None')
                    else:
                        try:
                            final_code_pcs, final_input_pcs = main_icd10_pcs(match)
                            final_code_pcs = clean_list(final_code_pcs)
                            if final_code_pcs:
                                for ji in final_code_pcs:
                                    Z_Result_code_list_first_match = []
                                    Result_code_list_first_match.append(ji)
                        except:
                            print("match", match)
                except:
                    time.sleep(5)
                    if code_list_first_match:
                        for i in range(len(code_list_first_match)):
                            if not str(code_list_first_match[i]).lower().startswith('w') and not str(code_list_first_match[i]).lower().startswith('x') and not str(code_list_first_match[i]).lower().startswith('y') and not str(code_list_first_match[i]).lower().startswith('z'):
                                Result_code_list_first_match.append(code_list_first_match[i])
                                Z_Result_code_list_first_match = []
                            else:
                                try:
                                    final_code, final_input = main_icd10_pcs(main_words_first_match[i])
                                    for j in final_code:
                                        Z_Result_code_list_first_match = []
                                        Result_code_list_first_match.append(j)
                                except:
                                    try:
                                        if len(code_list_first_match) == 1:
                                            print('Len=1')
                                            Result_code_list_first_match.append(code_list_first_match[0])
                                        else:
                                            Z_Result_code_list_first_match.append(code_list_first_match[i])
                                    except:
                                        print('None')
                    else:
                        try:
                            final_code_pcs, final_input_pcs = main_icd10_pcs(match)
                            final_code_pcs = clean_list(final_code_pcs)
                            if final_code_pcs:
                                for ji in final_code_pcs:
                                    Z_Result_code_list_first_match = []
                                    Result_code_list_first_match.append(ji)
                        except:
                            print("match", match)
                Result_code_list_first_match = clean_list(Result_code_list_first_match)
                Z_Result_code_list_first_match = clean_list(Z_Result_code_list_first_match)
                if not Result_code_list_first_match:
                    if Z_Result_code_list_first_match:
                        for ijf in Z_Result_code_list_first_match:
                            Result_code_list_first_match.append(ijf)
                for code in Result_code_list_first_match:
                    T_m = False
                    for Pro_code in Result:
                        if Pro_code in code:
                            T_m = True
                            code_matche_list.append(Result_code_list_first_match)
                            code_matches_Name.append(match)
                            break
                    if T_m:
                        break
                print("code_matche_list", code_matche_list)
                print('code_matches_Name', code_matches_Name)
        print('code_matche_list', code_matche_list)
        print('code_matches_Name', code_matches_Name)
        code_matche_list = clean_list(code_matche_list)
        print('code_matches_Name',code_matches_Name)
        print('code_matche_list', code_matche_list)
        if code_matche_list:
            TepM_out_list=[]
            matches = process.extract(procedure_name, code_matches_Name, scorer=fuzz.ratio, limit=5)
            final_Temp_match = [match[0] for match in matches]
            print('final_Temp_match',final_Temp_match)
            for match in final_Temp_match:
                for i in range(len(code_matches_Name)):
                    if str(code_matches_Name[i]).lower()==str(match).lower():
                        TepM_out_list.append(code_matche_list[i])
                        break
            print("matches", matches)
            print('TepM_out_list',TepM_out_list)
            temp_str = ""
            temp_percent = 0
            te_Code = []
            for match, percent in matches:
                possiblities.append(match)
                if percent >= temp_percent:
                    temp_str = match
                    if percent == 100:
                        temp_percent = percent
                        break
                    temp_percent = percent
            if temp_str == "":
                temp_str = matches[0][0]
            if temp_percent > 85:
                possiblities = []
            print(temp_str)
            for i in range(len(final_Temp_match)):
                if final_Temp_match[i] in temp_str:
                    for jj in range(len(TepM_out_list[i])):
                        tem_t = False
                        if not Master_id:
                            tem_t = True
                            Master_id = str(TepM_out_list[i][jj])
                        if Master_id and not tem_t:
                            Master_id = Master_id + ', ' + str(TepM_out_list[i][jj])
            master_index = packages_list.index(temp_str)
            master_index = int(package_indexes[master_index])
            print("master_index", master_index)
            print('------------End-----------------------')
        else:
            inter_matches = process.extract(procedure_name, first_match, scorer=fuzz.partial_ratio, limit=15)
            print("inter_matches", inter_matches)
            temp_list = procedure_name.lower().replace("(", "").replace(")", "").strip().split(" ")
            if "per" in temp_list:
                temp_list.remove("per")
            if "check" in temp_list:
                temp_list.remove("check")
            print("templist", temp_list)
            temp_match_list = []
            for match, percent in inter_matches:
                if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                    temp_match_list.append(match)
            if len(temp_match_list) == 0:
                temp_match_list = [match[0] for match in inter_matches]
            print("temp_match_list", temp_match_list)
            matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.ratio, limit=10)
            print('matches', matches)
            temp_match_list = []
            for match, percent in matches:
                if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                    temp_match_list.append(match)
            if len(temp_match_list) == 0:
                temp_match_list = [match[0] for match in matches]

            matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.partial_ratio, limit=5)

            print("matchesaasaaa", matches)
            if matches[0][1] <= 70:
                matches = process.extract(procedure_name, [match[0] for match in matches], scorer=fuzz.partial_ratio,
                                          limit=4)
            temp_str = ""
            temp_percent = 0
            for match, percent in matches:
                possiblities.append(match)
                if percent > temp_percent:
                    temp_str = match
                    # if len(match)/len(str(row["Name of the procedure/treatment"]).lower().replace("\n", "")) < 40.0:
                    #     continue
                    if percent == 100:
                        temp_percent = percent
                        break

                    temp_percent = percent
            if temp_str == "":
                temp_str = matches[0][0]
            if temp_percent > 85:
                possiblities = []
            print(temp_str)
            master_index = packages_list.index(temp_str)
            master_index = int(package_indexes[master_index])
            print(master_index)
        for room in rooms:
            print('JHgasuydjwewuyhwegvhew')
            ws.cell(row=KM, column=1).value = "FARIDABAD MEDICAL CENTRE"
            ws.cell(row=KM, column=2).value = ""
            ws.cell(row=KM, column=3).value = row["DESCRIPTION"]
            ws.cell(row=KM, column=4).value = id_codes
            ws.cell(row=KM, column=5).value = packages_items_df["ITEM_CODE"][master_index]
            ws.cell(row=KM, column=6).value = packages_items_df["NAME"][master_index]
            ws.cell(row=KM, column=7).value = ""
            ws.cell(row=KM, column=8).value = ""
            ws.cell(row=KM, column=9).value = rooms_dict[room]
            ws.cell(row=KM, column=10).value = row[room.replace('General ward', 'ECONOMY')]
            ws.cell(row=KM, column=11).value = ",".join(possiblities[1:])
            ws.cell(row=KM, column=12).value = Master_id
            if bool(",".join(possiblities[1:])):
                ws.cell(row=KM, column=11).fill = yellow_fill
            print(possiblities)
            KM += 1
    wb.save(r"VULVA AND VAGINA.xlsx")
    wb.close()
get_data()







