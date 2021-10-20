import time
from datetime import datetime
from googlesearch import search
import fuzzywuzzy.utils
import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill



def get_thevalues():
    excel_sheet = r"D:\Tariff Matching Master\HS00017509-Rate List\List.xlsx"
    master_file_name = r'D:\Tariff Matching Master\Reference\BIlling Code Master.xlsx'
    mappings_fileName = r'D:\Tariff Matching Master\Reference\mappings.xlsx'
    rooms = ["General ward", "SEMI PRIVATE", "PRIVATE", "DELUXE"]
    rooms_dict = {}
    st = datetime.now()

    master_df: pd.DataFrame = pd.read_excel(master_file_name, sheet_name="IRDA Updated",
                                            converters={'SUB_CATEGORY_CODE': str, 'CATEGORY_CODE': str})

    master_df = master_df.fillna("")
    master_df = master_df.replace(to_replace="NaN", value="")
    print((datetime.now() - st).total_seconds())
    print(master_df.keys())

    temp_room_df = master_df[master_df["SUB_CATEGORY_CODE"].str.contains("101000", na=False)]  # roomcharges
    print("temp_room_df", temp_room_df)
    names_list = temp_room_df["NAME"].str.lower().tolist()
    print("names_list", names_list)
    for room in rooms:
        print(room)
        matches = process.extractOne(room.lower(), names_list,
                                     scorer=fuzz.partial_ratio)
        print(matches)
        index = names_list.index(matches[0])
        print(index)
        rooms_dict[room] = temp_room_df["ITEM_CODE"][index + 1]

    print("rooms_dict", rooms_dict)

    inter_tariff_df = pd.DataFrame()
    list1 = []
    sheet_names_ = []
    sheet_names_1 = []
    for sheet in pd.ExcelFile(excel_sheet).sheet_names:
        packages_items_dfs: pd.DataFrame = master_df
        packages_list = packages_items_dfs["NAME"].str.lower().tolist()
        package_indexes = packages_items_dfs.index.tolist()
        print("names_list", packages_list)
        for room in packages_list:
            print(sheet.lower())
            if str(room).lower().__contains__(str(sheet).lower()):
                list1.append(room)
                sheet_names_.append(sheet)
                break
    for sheet in pd.ExcelFile(excel_sheet).sheet_names:
        if sheet not in sheet_names_:
            inter_tariff_df = inter_tariff_df.append(pd.read_excel(excel_sheet, sheet_name=sheet))
            sheet_names_1.append(sheet)

    wb = load_workbook(r"D:\Tariff Matching Master\HS00017509-Rate List\template.xlsx")
    j = 2
    yellow_fill = PatternFill(start_color="feff00",
                              fill_type='solid')
    if list1:
        for i in range(len(list1)):
            ws1 = wb.create_sheet(list1[i])
            master_index = packages_list.index(list1[i])
            master_index = int(package_indexes[master_index])
            print("master_index", master_index)
            count = packages_items_dfs["SUB_CATEGORY_CODE"][master_index]
            na = packages_items_dfs["NAME"][master_index]
            print("count", count)
            print("na", na)
            Main_frames_to_check = master_df[master_df["SUB_CATEGORY_CODE"].str.contains(count, na=False)]#main

            inter_tariff_dfss = pd.DataFrame();
            inter_tariff_dfss = inter_tariff_dfss.append(pd.read_excel(excel_sheet, sheet_name=sheet_names_[i]))
            print(inter_tariff_dfss)
            inter_tariff_dfss = inter_tariff_dfss.fillna("")
            inter_tariff_dfss = inter_tariff_dfss.replace("NA", "0.0")
            services_lists = inter_tariff_dfss["DESCRIPTION"].tolist()

            packages_items_dfs: pd.DataFrame = Main_frames_to_check
            packages_listss = packages_items_dfs["NAME"].str.lower().tolist()
            package_indexessss = packages_items_dfs.index.tolist()
            print("package_indexessss",packages_listss)
            print('package_indexessss',package_indexessss)

            mappings_dfs = pd.read_excel(mappings_fileName, sheet_name="Sheet1", converters={'Name': str})
            mappings_list1 = mappings_dfs["Name"].str.lower().tolist()

            print(type(mappings_list1))

            for index, row in inter_tariff_dfss.iterrows():
                id_code=''
                if not bool(str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")):
                    continue
                possiblities = []

                procedure_name = str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")
                print(type(procedure_name))
                i, z = process.extractOne(procedure_name, mappings_list1, scorer=fuzz.ratio)
                print("i", i)
                print("z", z)

                if z > 95:
                    procedure_name = mappings_dfs["Possibility"][mappings_list1.index(i)]
                print("actual", str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " "))

                procedure_name = procedure_name.replace("therapeutic", "diagnostic/therapeutic").replace("charges",
                                                                                                         "charge")
                print("procedure", procedure_name)
                id_code=icd10_code(procedure_name)
                matches = process.extract(procedure_name, packages_listss, scorer=fuzz.partial_token_sort_ratio, limit=30)
                print("first", matches)
                first_match = [match[0] for match in matches]
                print('first_match', first_match)
                inter_matches = process.extract(procedure_name, first_match, scorer=fuzz.partial_ratio, limit=15)
                print("inter_matches", inter_matches)
                temp_list = procedure_name.lower().replace("(", "").replace(")", "").strip().split(" ")
                if "per" in temp_list:
                    temp_list.remove("per")
                if "check" in temp_list:
                    temp_list.remove("check")
                print("templist", temp_list)
                temp_match_list = []
                for match, percent in inter_matches:
                    if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                        temp_match_list.append(match)
                if len(temp_match_list) == 0:
                    temp_match_list = [match[0] for match in inter_matches]
                print("temp_match_list", temp_match_list)
                matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.ratio, limit=10)
                print('matches', matches)
                temp_match_list = []
                for match, percent in matches:
                    if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                        temp_match_list.append(match)
                if len(temp_match_list) == 0:
                    temp_match_list = [match[0] for match in matches]

                matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.partial_ratio, limit=5)

                print("matchesaasaaa", matches)
                if matches[0][1] <= 70:
                    matches = process.extract(procedure_name, [match[0] for match in matches],
                                              scorer=fuzz.partial_ratio,
                                              limit=4)
                temp_str = ""
                temp_percent = 0
                for match, percent in matches:
                    possiblities.append(match)
                    if percent > temp_percent:
                        temp_str = match
                        # if len(match)/len(str(row["Name of the procedure/treatment"]).lower().replace("\n", "")) < 40.0:
                        #     continue
                        if percent == 100:
                            temp_percent = percent
                            break

                        temp_percent = percent
                if temp_str == "":
                    temp_str = matches[0][0]
                if temp_percent > 85:
                    possiblities = []
                print(temp_str)
                master_index = packages_listss.index(temp_str)
                master_index = int(package_indexessss[master_index])
                print(master_index)
                for room in rooms:
                    ws1.cell(row=j, column=1).value = "FARIDABAD MEDICAL CENTRE"
                    ws1.cell(row=j, column=2).value = ""
                    ws1.cell(row=j, column=3).value = row["DESCRIPTION"]
                    ws1.cell(row=j, column=4).value = id_code
                    ws1.cell(row=j, column=5).value = packages_items_dfs["ITEM_CODE"][master_index]
                    ws1.cell(row=j, column=6).value = packages_items_dfs["NAME"][master_index]
                    ws1.cell(row=j, column=7).value = ""
                    ws1.cell(row=j, column=8).value = ""
                    ws1.cell(row=j, column=9).value = rooms_dict[room]
                    ws1.cell(row=j, column=10).value = row[room.replace('General ward', 'ECONOMY')]
                    ws1.cell(row=j, column=11).value = ",".join(possiblities[1:])
                    if bool(",".join(possiblities[1:])):
                        ws1.cell(row=j, column=11).fill = yellow_fill
                    print(possiblities)
                    j += 1


    inter_tariff_df = inter_tariff_df.fillna("")
    inter_tariff_df = inter_tariff_df.replace("NA", "0.0")
    inter_tariff_df.to_excel("temp.xlsx")
    services_list = inter_tariff_df["DESCRIPTION"].tolist()
    print("services_list", services_list)
    packages_items_df: pd.DataFrame = master_df
    packages_list = packages_items_df["NAME"].str.lower().tolist()
    package_indexes = packages_items_df.index.tolist()
    print("packages_list", packages_list)
    print("package_indexes", package_indexes)
    mappings_df = pd.read_excel(mappings_fileName, sheet_name="Sheet1", converters={'Name': str})
    mappings_list = mappings_df["Name"].str.lower().tolist()

    print(type(mappings_list))
    ws2 = wb.create_sheet('L1')
    j = 2
    for index, row in inter_tariff_df.iterrows():
        id_codes = ''
        if not bool(str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")):
            continue
        possiblities = []

        procedure_name = str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " ")
        print(type(procedure_name))
        i, z = process.extractOne(procedure_name, mappings_list, scorer=fuzz.ratio)
        print("i", i)
        print("z", z)

        if z > 95:
            procedure_name = mappings_df["Possibility"][mappings_list.index(i)]
        print("actual", str(row["DESCRIPTION"]).lower().replace("\r", " ").replace("\n", " "))

        procedure_name = procedure_name.replace("therapeutic", "diagnostic/therapeutic").replace("charges", "charge")
        print("procedure", procedure_name)
        id_codes = icd10_code(procedure_name)
        matches = process.extract(procedure_name, packages_list, scorer=fuzz.partial_token_sort_ratio, limit=30)
        print("first", matches)
        first_match = [match[0] for match in matches]
        print('first_match', first_match)
        inter_matches = process.extract(procedure_name, first_match, scorer=fuzz.partial_ratio, limit=15)
        print("inter_matches", inter_matches)
        temp_list = procedure_name.lower().replace("(", "").replace(")", "").strip().split(" ")
        if "per" in temp_list:
            temp_list.remove("per")
        if "check" in temp_list:
            temp_list.remove("check")
        print("templist", temp_list)
        temp_match_list = []
        for match, percent in inter_matches:
            if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                temp_match_list.append(match)
        if len(temp_match_list) == 0:
            temp_match_list = [match[0] for match in inter_matches]
        print("temp_match_list", temp_match_list)
        matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.ratio, limit=10)
        print('matches', matches)
        temp_match_list = []
        for match, percent in matches:
            if temp_list[0] in match and temp_list[len(temp_list) - 1] in match:
                temp_match_list.append(match)
        if len(temp_match_list) == 0:
            temp_match_list = [match[0] for match in matches]

        matches = process.extract(procedure_name, temp_match_list, scorer=fuzz.partial_ratio, limit=5)

        print("matchesaasaaa", matches)
        if matches[0][1] <= 70:
            matches = process.extract(procedure_name, [match[0] for match in matches], scorer=fuzz.partial_ratio,
                                      limit=4)
        temp_str = ""
        temp_percent = 0
        for match, percent in matches:
            possiblities.append(match)
            if percent > temp_percent:
                temp_str = match
                # if len(match)/len(str(row["Name of the procedure/treatment"]).lower().replace("\n", "")) < 40.0:
                #     continue
                if percent == 100:
                    temp_percent = percent
                    break

                temp_percent = percent
        if temp_str == "":
            temp_str = matches[0][0]
        if temp_percent > 85:
            possiblities = []
        print(temp_str)
        master_index = packages_list.index(temp_str)
        master_index = int(package_indexes[master_index])
        print(master_index)
        print(packages_items_df["ITEM_CODE"][4265])
        for room in rooms:
            print('JHgasuydjwewuyhwegvhew')
            ws2.cell(row=j, column=1).value = "FARIDABAD MEDICAL CENTRE"
            ws2.cell(row=j, column=2).value = ""
            ws2.cell(row=j, column=3).value = row["DESCRIPTION"]
            ws2.cell(row=j, column=4).value = id_codes
            ws2.cell(row=j, column=5).value = packages_items_df["ITEM_CODE"][master_index]
            ws2.cell(row=j, column=6).value = packages_items_df["NAME"][master_index]
            ws2.cell(row=j, column=7).value = ""
            ws2.cell(row=j, column=8).value = ""
            ws2.cell(row=j, column=9).value = rooms_dict[room]
            ws2.cell(row=j, column=10).value = row[room.replace('General ward', 'ECONOMY')]
            ws2.cell(row=j, column=11).value = ",".join(possiblities[1:])
            if bool(",".join(possiblities[1:])):
                ws2.cell(row=j, column=11).fill = yellow_fill
            print(possiblities)
            j += 1
    wb.save(r"FARIDABAD_MEDICAL_CENTRE_OUTPUT.xlsx")
    wb.close()


def icd10_code(dieases_data):
    # dieases_data = input("diagnosis name : ")
    query = str(dieases_data) + " icd 10"
    list_url = []
    for i in search(query, lang='en'):
        list_url.append(i)
    for i in range(len(list_url)):
        if str(list_url[i]).__contains__('icd10data'):
        # print(list_url[i])
            my_split = str(list_url[i]).split('/')[-1]
            return my_split

if __name__ == "__main__":
    get_thevalues()



























