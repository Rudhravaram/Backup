import time
from datetime import datetime
from googlesearch import search
import fuzzywuzzy.utils
import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_thevalues():
    master_file_name = r'D:\Teams extracted data\Master_extreat\Reference\BIlling Code Master.xlsx'
    rooms_dict = {}
    st = datetime.now()
    master_df: pd.DataFrame = pd.read_excel(master_file_name, sheet_name="Sheet3",
                                            converters={'SUB_CATEGORY_CODE': str, 'CATEGORY_CODE': str})
    master_df = master_df.fillna("")
    master_df = master_df.replace(to_replace="NaN", value="")
    print((datetime.now() - st).total_seconds())
    print(master_df.keys())
    packages_items_df: pd.DataFrame = master_df
    packages_list = packages_items_df["NAME"].str.lower().tolist()
    package_indexes = packages_items_df.index.tolist()
    wb = load_workbook(master_file_name)
    ws = wb["Sheet3"]
    j = 2
    for i in packages_list:
        print(i)
        Result=icd10_code(i)
        master_index = packages_list.index(i)
        master_index = int(package_indexes[master_index])
        ws.cell(row=j, column=5).value = Result
        j += 1
    wb.save(r"dev.xlsx")
    wb.close()

def icd10_code(dieases_data):
    # dieases_data = input("diagnosis name : ")
    query = str(dieases_data) + " icd 10"
    list_url = []
    for i in search(query, lang='en'):
        list_url.append(i)
    for i in range(len(list_url)):
        if str(list_url[i]).__contains__('icd10data'):
            # print(list_url[i])
            my_split = str(list_url[i]).split('/')[-1]
            return my_split
if __name__ == "__main__":
    get_thevalues()