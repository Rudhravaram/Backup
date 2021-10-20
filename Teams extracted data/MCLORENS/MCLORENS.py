import os
import PyPDF2
import re
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import openpyxl
import csv

# files=['IAR-1806-18024_Reliance Retail Ltd_Bihar_FSR_259371167.txt','IAR-1805-17754_Reliance Retail Ltd_Kolhapur_FSR_255404966.txt','IAR-1805-17556_Reliance Retail Ltd_Kerala_FSR_255225729.txt','IAR-1804-16213_Reliance Retail Ltd_Chhatisgarh_FSR_255270968.txt','fsr_249640647.txt','fsr_249926810.txt','FSR_261474078.txt','FSR_260971977.txt','FSR_260103398.txt','fsr_260006564.txt','FSR_253219694.txt','FSR_253171607.txt','FSR_253181884.txt','FSR_253235102.txt','FSR_253235108.txt','FSR_254382974.txt','FSR_254404042.txt','FSR_255076580.txt','FSR_255137074.txt','FSR_255223948.txt','FSR_255228497.txt','FSR_255356044.txt','FSR_257497139.txt','FSR_258871989.txt','FSR_259642277.txt','FSR_259928215.txt']
# files=['IAR-2007-11314-Final report_423874757.txt','IAR-1806-18967_Reliance Retail Limited_Dehradun_FSR_261476517.txt','IAR-1806-18024_Reliance Retail Ltd_Bihar_FSR_259377850.txt']
# files=['fsr_249640647.txt','10451_FIR029215418_reliane retail_FSR_408129929.txt','10437_FIR027666496_Reliance Retail_FSR_408278765.txt','10435_FIR027655130_Reliance Retail_FSR_408231507.txt','10433_FIR027639039_Reliance Retail_FSR_408220587.txt','IAR-1806-18967_Reliance Retail Limited_Dehradun_FSR_261476517.txt','FSR_261474078.txt','IAR-1806-18024_Reliance Retail Ltd_Bihar_FSR_259377850.txt','IAR-1806-18024_Reliance Retail Ltd_Bihar_FSR_259371167.txt','IAR-1805-17754_Reliance Retail Ltd_Kolhapur_FSR_255404966.txt','IAR-1805-17556_Reliance Retail Ltd_Kerala_FSR_255225729.txt','IAR-1804-16213_Reliance Retail Ltd_Chhatisgarh_FSR_255270968.txt','FSR_259642277.txt', 'FSR_259928215.txt','FSR_255223948.txt', 'FSR_255228497.txt', 'FSR_255356044.txt', 'FSR_257497139.txt', 'FSR_258871989.txt','FSR_253235108.txt', 'FSR_254382974.txt', 'FSR_254404042.txt', 'FSR_255076580.txt', 'FSR_255137074.txt','fsr_260006564.txt', 'FSR_253219694.txt', 'FSR_253171607.txt', 'FSR_253181884.txt', 'FSR_253235102.txt','fsr_249640647.txt', 'fsr_249926810.txt', 'FSR_261474078.txt', 'FSR_260971977.txt', 'FSR_260103398.txt','IAR-2009-42597-FSR_425067822.txt','IAR-2008-42405-FSR_421724782.txt','IAR-2009-42999-FSR_425462510.txt','IAR-2009-42734-FSR_424072739.txt','IAR-2009-42733-FSR_424076180.txt','IAR-2009-42731-FSR_424065681.txt','IAR-2008-42380_421272738.txt','IAR-2008-42384-FSR_421282500.txt','IAR-2009-43191-FSR_427617271.txt','IAR-2009-42825-FSR_425184604.txt','IAR-2009-42599-FSR_425085530.txt','IAR-2009-42597-FSR_425069886.txt','IAR-2009-42999-FSR_425462510.txt','IAR-2008-42405-FSR_421724782.txt','IAR-2009-42597-FSR_425067822.txt','IAR-2009-42597-FSR_425069886.txt']
# files=['10433_FIR027639039_Reliance Retail_FSR_408220587.txt']
files = [each for each in os.listdir(r'C:\Teams extracted data\MCLORENS') if each.endswith('.txt')]
path = 'IAR.xlsx'
cnt = 0
# ///before loop starts
filename1 = "Excel_Files.csv"
with open(filename1, newline='', encoding='utf-8', errors='ignore') as f:
    reader = csv.reader(f)
    data1 = list(reader)
data = []
cn = 0
for elem in data1[1:]:
    string1 = ''
    for item2 in elem:
        string1 = string1 + item2 + '&$'
    data.append(string1)
# print(data)
filename2 = "NIL Claim.csv"
with open(filename2, newline='', encoding='utf-8', errors='ignore') as f:
    reader = csv.reader(f)
    data1 = list(reader)
data_one = []
cn = 0
for elem in data1[1:]:
    string1 = ''
    for item2 in elem:
        string1 = string1 + item2 + '&$'
    data_one.append(string1)


# print(data_one)

def city_finder(txt1):
    pin = ''
    p = []
    for t in txt1.split(','):
        t = re.sub(' +', '', t)
        p = re.findall('(?<=[^_])\d{6}', t)
        if len(p) > 0 and len(p[0]) == 6:
            pin = p[0]
    txt = []
    dis_lis = []
    for t in txt1.split(','):
        t = re.sub('[\d+]', '', t)
        txt.append(t.title().replace('.', '').replace('-', '').replace('/', '').strip())
    dis_lis = txt[-3:]
    district = ''
    dist_list = []
    filename1 = 'district.csv'
    with open(filename1, newline='', encoding='utf-8', errors='ignore') as f:
        reader = csv.reader(f)
        data1 = list(reader)
        data = []
        for elem in data1:
            data.append(elem[0])
        dist_list = data
        for elem in data:
            for dl in dis_lis:
                if fuzz.ratio(dl, elem) > 90:
                    district = elem
    if len(district) == 0 and len(pin) != 0:
        filename2 = 'pin_code.csv'
        with open(filename2, newline='', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            data1 = list(reader)
        data = []
        for elem in data1:
            string1 = ''
            for item2 in elem:
                string1 = string1 + item2 + '$&'
            data.append(string1)
        pinV = ''
        for item1 in data:
            if pin == item1.split('$&')[1]:
                district = item1.split('$&')[8]

    elif len(district) == 0 and len(pin) == 0:
        if (',') in txt1:
            for tx in txt1.split(',')[::-1]:
                tx = tx.replace('-', ' ').title().strip()
                for t in tx.split():
                    t = t.replace('-', ' ').title().strip()
                    # print('tttt',t)
                    for d in dist_list:
                        if t.strip() == d.strip():
                            district = t
                            break
    # print(pin)
    return district


dirs = os.listdir(r'C:\Teams extracted data\MCLORENS')
for fl in dirs:
    if fl.endswith(".txt"):
        txt = ''
        with open(fl, mode='rt', encoding='utf-8') as f:
            for l in f:
                # print(l)
                txt += l + ' '
        #     file = open(filename, mode='rt', encoding='utf-16-le')
        # pdfFileObj = open(fl, 'rb')
        # pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        # n=	pdfReader.numPages
        # txt=''
        # with open (fl.split('.')[0]+'.txt','w') as f:
        #     for i in range(n):
        #         pageObj = pdfReader.getPage(i)
        #         print(pageObj.extractText())
        #         f.write(pageObj.extractText())
        #         for w in pageObj.extractText().split('\n'):
        #             if len(w.replace(' ',''))==0:
        #                 continue
        #             txt +=w +' '
        # pdfFileObj.close()
        # f.close()
        #     txt=open(txt, mode='rt', encoding='utf-16-le')
        txt = re.sub(' +', ' ', txt)
        txt = txt.replace('\n', '')
        txt = txt.replace('\f', '')
        # 'Excess':{'ADJUSTMENT','BACKGROUND & OCCURRENCE','BACKGROUND AND OCCURRENCE','being deducted','for each','NET ADJUSTED','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss','NET ADJUSTED CLAIM'},'EXCESS':{'NET ADJUSTED CLAIM','ADJUSTMENT','NET ADJUSTED','NET ADJUSTED CLAIM','BACKGROUND & OCCURRENCE:','ADJUSTMENT','being deducted'},
        # print(txt)
        #     dic_key={'IL Claim Ref No':{'CLAIM NO.':['Team Ref. No.','Date:']},
        #             'Date of Loss':{'TIME & DATE OF LOSS':{'CIRCUMSTANCES OF'},'TIME, DAY, DATE OF':{'SITUATION OF LOSS'},'TIME, DAY & DATE OF':{'THE INSURED'},'DATE OF LOSS':{'SITUATION OF LOSS'}},
        #             'Cause of Loss':{'APPARENT CAUSE':{'NATURE & APPARENT'},'ROOT CAUSE':{'NATURE, EXTENT OF'},'CAUSE OF LOSS':{'NATURE, EXTENT OF'}},
        #             'Cause of Loss (Sentence) Description':{'CIRCUMSTANCES OF':['APPARENT CAUSE','NATURE AND EXTENT OF DAMAGE']},
        #             'Estimate of Loss':{'FINAL CLAIM':['COST OF RESTORATION','LOSS ASSESS']},
        #             'Loss Location (Full) Address':{'RISK LOCATION':{'PROPERTY INSURED'},'LOCATION OF LOSS':{'Corporate'},'DATE AND PLACE OF':{'GROSS ASSESSED LOSS'},'LOCATION COVERED':{'Corporate'}},
        #             'Date of Survey':{'DATE AND PLACE OF':{'Corporate'}},
        #             'Final Amount':{'CONCURRENCE':{'LOSS'}},
        #             'Gross Loss':{'GROSS ASSESSED LOSS':['NET ADJUSTED LOSS','SETTLEMENT'],'DEDUCTIBLE':{'LOSS'}},
        #             'Salvage':{'SALVAGE CONSIDERED':{'NET LOSS ASSESSED'},'ADEQUACY OF INSURANCE':{'SALVAGE'}},
        #             'Under Insurance':{'UNDER INSURANCE':{'DEDUCTIBLE'}},
        #             'Excess':{'LESS: EXCESS':{'NET ADJUSTED LOSS'},'POLICY EXCESS':{'ADJUSTMENT AS PER POLICY'}},
        #             'Depreciation':{'LESS: DEPRICIATION':{'MARKET VALUE OF LOSS'}},
        #             'Assessment (Full)':{'ITEM OF LOSS, AS PER':{'SUM INSURED'}},
        #             'Property Damaged':{'SL DESCRIPTION QTY AMOUNT (RS)':{'GROSS ASSESSED LOSS'}},
        #             'Amount per Quantity':{'SL DESCRIPTION QTY AMOUNT (RS)':{'GROSS ASSESSED LOSS'}}lclclLmb
        #             }
        # clm_no=((txt.split(k2)[1]).split(x)[0]).strip()CLAIM NUMBER 'SALVAGE':{}
        dic_key = {'IL Claim Ref No': {'CLAIM REFERENCE': {'MCLARENS REFERENCE'},
                                       'Green house claims': {'attachment', '[External'},
                                       'claim under IAR Policy': {'Reserve'}, 'ICICI Claim Ref# ': {'Insured'},
                                       'Claim RefNo': 'POLICY', 'Claim RclNo': {'POLICY', 'M/s'},
                                       'Claim Ref No': ['Dear', 'INSURED'],
                                       'Claim RcfNo.': 'POLICY',
                                       'CLAIM REF No': 'In',
                                       'CLAIM REF NO': 'In',
                                       'Claim Number': {'Date of Loss'},
                                       'CLAIM NUMBER': 'PERIOD OF',
                                       'CLAIM NO': {'INSTRUCTIONS', 'POLICY NO.'},
                                       'Underwriters': {'No.'},
                                       'Claim No': [')', 'Date', 'ICICI', 'lClCl', 'Utsav', 'Policy', 'Submitted',
                                                    'Pursuant', 'FINAL SURVEY'],
                                       'Claim #': 'Policy #', '[ CLAIM REF. NO:': ']',
                                       '[CLAIM NO': ']',
                                       '( Claim No': {')'}, 'clai.n No': {'lclclLmb', 'lClCl', 'ICICI'},
                                       'Chim No': {'ICICI', 'lClCl'}, 'ClaimNo': 'ICICI',
                                       'ClaiD No': 'WITHOUT',
                                       'Claim Num ber': {'Date'},
                                       'CLAIM REF. NO': {'TRADE'}
                                       },
                   'Date of Loss': {'assessment of loss vide mail dated': {'EXTENT OF'},
                                    'DATE OF LOSS': {'DATE OF INSTRUCTION', 'Regd.', 'THE INCIDENT', 'CLAIM INTIMATION',
                                                     'Regd. Office', '3.02 NATURE', 'DATE OF INTIMATION',
                                                     'CAUSE OF LOSS', 'NATURE OF LOSS', 'CLAIM INTIMATION'},
                                    'DATES OF LOSS': {'CAUSE OF LOSS'},
                                    'Date of Loss': {'Allotment Date', 'Affected WTG', 'Location', ', when', 'the',
                                                     'Date of Intimation', 'DATE OF INTIMATION', 'Date of Request'},
                                    'Date & Time of Loss': {'at'}, 'Date of loss': {'THE', 'Loss', 'Location of loss'},
                                    'DATE & TIME OF LOSS': {'DATE'}, 'LOSS DETAILS': 'SURVEY',
                                    'Date of reported Breakdown of Gear Box of ZR02 WTG:': 'Date of Intimation',
                                    'Sub-station at Akal Village on': '. The', 'DOL': ', Claim',
                                    'Ambala Chandigarh Expressways on': 'at',
                                    'TIME, DAY & DATE OF': {', as reported', 'INCIDENT'},
                                    'TIME, DAY & DATE OF LOSS': ', as reported',
                                    'Date of Damage': 'Date of Appointment',
                                    'Incident': 'Pipe/equipment',
                                    'Date of Breakdown': 'Dear Sir',
                                    'occurred due to Gaja Cyclone on': 'With regard',
                                    'delivering 157 MT of steam. On': {'morning around'},
                                    'all due to water logging on': {'at N4/s Kohinoor'},
                                    'Builcling and P&N4 due to Inundation on': {'at NI/s Hotel'},
                                    'all due to accident on': {'at H. No. 302'},
                                    'Pre-Grinding Mill (VRPM) got damaged on': {'early morning'},
                                    'OATE OF LOSS': {'DATE OF II', 'DATE'},
                                    'Date o{ loss': {'Loss site'},
                                    'reported to have taken place on at': {'Earlier, Boiler #1 had'},
                                    'oaTE OF  LOSS': {'oarE oF lnflIAtior'}
                                    },
                   'Cause of Loss': {
                       'LOSS / DAMAGE': {'RECOVERY'},
                       'CAUSE': {'NATURE AND EXTENT OF'},
                       'OPIMON ON LOSS': {'EXTENT OF LOSS'},
                       'OPINION ON LOSS': {'EXTENT OF LOSS'},
                       'CAUSE OF BREAKDOWN': {'INSURED’S'},
                       'THE INCIDENT': {'INSURANCE', 'POLICY NO.'},
                       'Loss elescription': {'i/We'},
                       'ON A-lC OF DAMAGE': {'ON 02-04.11.19'},
                       'a./c ofdamage to': {'on 15.09.19'},
                       'CAUSE OF LOSS:': {'Offices:', 'Estimated Loss', 'Loss Address', 'SALVAGE:', 'COVERAGE',
                                          'DOCUMENTS', 'NATURE', 'LIABILITY', 'DOCUMENTS EXAMINED',
                                          'COVERAGE & ADMISSIBILITY', 'NATURE & EXTENT OF'},
                       'Cause of Loss': {'Loss Address1', 'SALVAGE:', 'COVERAGE', 'DOCUMENTS', 'NATURE', 'LIABILITY',
                                         'Estimated Loss', 'Loss Address', 'Adequacy of Sum', 'NATURE & EXTENT OF'},
                       'EVENT': {'INTRODUCTION'},
                       'Type of Loss': {'Material'},
                       'SITUATION OF LOSS': {'Regd'},
                       'Cause of loss': {'This'},
                       'Cause of the loss': {'Liability'},
                       'CAUSE OF DAMAGE:': {'ADMISSIBILITY'},
                       'APPARENT CAUSE': {'NATURE & APPARENT'}
                   },
                   'Cause of Loss (Sentence) Description': {
                       #                                                         'SITUATION OF LOSS':{'ADDRESS OF LOSS'},
                       #                                                         'NATURE AND EXTENT OF DAMAGE':{'PAYMENT RECOMMENDED'},
                       'LOSS / DAMAGE': {'RECOVERY'},
                       'CAUSE': {'NATURE AND EXTENT OF'},
                       'OPIMON ON LOSS': {'EXTENT OF LOSS'},
                       'OPINION ON LOSS': {'EXTENT OF LOSS'},
                       'CAUSE OF BREAKDOWN': {'INSURED’S'},
                       'THE INCIDENT': {'INSURANCE', 'POLICY NO.'},
                       'Loss elescription': {'i/We'},
                       'ON A-lC OF DAMAGE': {'ON 02-04.11.19'},
                       'a./c ofdamage to': {'on 15.09.19'},
                       'CAUSE OF LOSS:': {'Offices:', 'Estimated Loss', 'Loss Address', 'SALVAGE:', 'COVERAGE',
                                          'DOCUMENTS', 'NATURE', 'LIABILITY', 'DOCUMENTS EXAMINED',
                                          'COVERAGE & ADMISSIBILITY', 'NATURE & EXTENT OF'},
                       'Cause of Loss': {'Loss Address1', 'SALVAGE:', 'COVERAGE', 'DOCUMENTS', 'NATURE', 'LIABILITY',
                                         'Estimated Loss', 'Loss Address', 'Adequacy of Sum', 'NATURE & EXTENT OF'},
                       'EVENT': {'INTRODUCTION'},
                       'Type of Loss': {'Material'},
                       'SITUATION OF LOSS': {'Regd'},
                       'Cause of loss': {'This'},
                       'Cause of the loss': {'Liability'},
                       'CAUSE OF DAMAGE:': {'ADMISSIBILITY'},
                       'APPARENT CAUSE': {'NATURE & APPARENT'},
                       'CIRCUMSTAI\CES OF INCIDENT': {'NATURE'},
                       'NATURE OF LOSS': {'3.03 AFFECTED LOCATION'}, 'CAUSE OF BREAKDOWN': {'INSURED’S'},
                       'THE INCIDENT': {'INSURANCE', 'POLICY NO.'},
                       'CAUSE OF LOSS:': {'Offices:', 'Estimated Loss', 'Loss Address', 'SALVAGE:', 'COVERAGE',
                                          'DOCUMENTS', 'NATURE', 'LIABILITY', 'DOCUMENTS EXAMINED',
                                          'COVERAGE & ADMISSIBILITY', 'NATURE & EXTENT OF', 'Page'},
                       'Cause of Loss': {'Loss Address1', 'SALVAGE:', 'NATURE & EXTENT OF', 'COVERAGE', 'DOCUMENTS',
                                         'NATURE', 'LIABILITY', 'Estimated Loss', 'Loss Address', 'Adequacy of Sum'},
                       'NATURE & EXTENT OF': {'LOSS RESERVE', 'INSURED’S'},
                       'CAUSE OF': {'LIABILITY'},
                       'EVENT': {'INTRODUCTION'},
                       'Type of Loss': 'Material',
                       'SITUATION OF LOSS': 'Regd',
                       'Loss elescription': 'i/We',
                       'Cause of loss': 'This',
                       'ON A-lC OF DAMAGE': 'ON 02-04.11.19',
                       'a./c ofdamage to': 'on 15.09.19',
                       'Cause of the loss': 'Liability',
                       'Cause of loss': 'Nature of',
                       'OCCURRENCE': {'INSPECTION'},
                       'BACKGROUND & OCCURRENCE:': {'OUR SURVEY:'},
                       'OCCURRENCE & CAUSE': {'EXTENT OF'},
                       'Cause of Loss': {'Adequacy of Sum'},
                       'LIABILITY UNDER': {'INSURED’S FINAL'},
                       'CIRCUMSTANCES OF INCIDENT': ['Regd.', '23rd March 2018', 'EVIDENCES OF OCCURRENCE',
                                                     'CAUSE OF LOSS', 'APPARENT CAUSE'],
                       'CAUSE OF LOSS': {'INSURED’S', 'Mumbai Office', 'SALVAGE'},
                       'CIRCUMSTANCES OF LOSS': {'Regd'},
                       #                                                         'SITUATION OF LOSS':{'Mumbai Office'},
                       'BACKGROUND': {'INSURED’S CLAIM'},
                       'Loss Description': {'Loss Date'},
                       'OPINION ON LOSS': {'EXTENT OF LOSS'},
                       'CAUSE OF DAMAGE:': {'ADMISSIBILITY'},
                       #                                                         'CIRCUMSTANCES':{'ESTIMATED AMOUNT'},
                       'OPIMON ON LOSS': {'EXTENT OF LOSS'},
                       'NATURE & EXTENT OF': {'LOSS RESERVE', 'INSURED’S'},
                       'NATURE & EXTENT OF DAMAGE': {'POLICY CONDITIONS', 'INSURED’S'},

                   },
                   'Estimate of Loss': {'BROADCAST LTD': {'THEFT'}, 'Estimated Loss': {'Assessed Loss'},
                                        'ESTIMATED LOSS': {'THE SOLAR PUMP HOUSE'},
                                        'INSURED’S CLAIM:-': {'ASSESSMENT OF LOSS', 'ADEQUACY OF INSURANCE'},
                                        'Assessment of loss': {'Page No'}, 'Total claimed Amount': {'LOSS ASSESSMENT'},
                                        'Total Claim Value': {'SALVAGE'},
                                        'CLAIM AMOUNT': ['GROSS LOSS', 'ASSESSMENT OF', 'LOSS ASSESSMENT'],
                                        'INSURED’S FINAL CLAIM': ['LOSS ASSESSMENT', 'DEDUCTIBLE',
                                                                  'INVOICE VALUE OF LOSS', 'LOSS ASSESS',
                                                                  'ASSESSMENT OF LOSS', 'GROSS LOSS'],
                                        'FINAL CLAIM': ['COST OF RESTORATION', 'LOSS ASSESS', 'INVOICE VALUE OF LOSS',
                                                        'GROSS LOSS'], 'AMOUNT CLAIMED': {'GROSS ASSESSED'},
                                        'INSURED’S CLAIM': ['Description', 'Particular', 'LOSS ASSESSMENT',
                                                            'ASSESSMENT', 'SALVAGE']},
                   'Loss Location (Full) Address': {'ADDRESS (LOSS LOCATION)': {'PERSON CONTACTED'},
                                                    'Risk Location / Address': {'6 Nature of Business'},
                                                    'AFFECTED LOCATION': {'4.00 SURVEY PARTICULARS'},
                                                    'Loss Site': {'Policy Ref.'},
                                                    'Risk Address': {'IAR-1804-16045_Reliance'},
                                                    'Risk Location': {'Section wise', 'Sum Insured'},
                                                    'RISK LOCATION': {'OCCUPATION OF'},
                                                    'ADDRESS OF LOSS': {'DATE OF LOSS', 'HYPOTHECATION'},
                                                    'PLACE OF SURVEY': {'REPRESENTATIVE'},
                                                    'LOSS LOCATION': {'DATE OF LOSS', '[0-li=E*'},
                                                    'LOSSLOCATIOII': {'oaTE OF  LOSS'}
                                                    },
                   'Date of Survey': {'survey was conducted by us on': {'After', 'The'},
                                      'SURVEY DATE': {'ADMISSIBILTY'},
                                      'survey conducted on': {'and'},
                                      'DATE OF VISIT': {'CAUSE OF LOSS', 'DESCRIPTION', ', 6th', 'REASON FOR DELAY',
                                                        'Office'},
                                      'Survey Date': {'Damage', 'Loss'},
                                      'DATE OF SURVEY': {'and subsequently', 'PERSON CONTACTED', 'CONTACT PERSONS',
                                                         'PLACE OF SURVEY'},
                                      'survey was conducted on': {'at'}, 'from the underwriters on': 'to survey',
                                      'Date of Survey': {'IP as', 'Affected Location', 'Telecom Site'},
                                      'claim was carried out on': 'During',
                                      'at Akal Village Jaipur on': ', in',
                                      'visited the plant on': 'and met',
                                      'loss location of Insured on dated': 'at Milkat',
                                      'DATE & PLACE OF SURVEY': 'at', 'Pradesh) on': 'to survey',
                                      'Mettur Dam, Salem on': "and carried",
                                      'Dates of Survey': 'and subsequent',
                                      'carried out by us on': 'During survey',
                                      'PS-Banarpal Dist Angul, Orissa on': 'to survey',
                                      'Cuttack Orissa on': 'to survey',
                                      'Dharkhola, Keonjhar Orissa on': 'to survey',
                                      'Vedaranyam Taluk, Nagapattinam – 614 810 on': ', 07th',
                                      'Date of survey': 'Cause of',
                                      'Tehsil,Gharghoda, Dist. Raigarh, Chhattisgarh – 496107 on': '. Mr.',
                                      'visited the loss location on': {', as per'},
                                      'We visited the factory for the 1st time on': {', accompanied'},
                                      'we visited': {'& on subsequent'},
                                      'surr.c,v was conducted by us on': {'and thc lnsr.rrcd'},
                                      'Accordingly, surve) \vas conclucted L.r-v Lrs oir': {', 10.08.19 &'},
                                      'Accordingly, Sutvey was conducted by us on': {'and the Insured rvas'},
                                      'Works, Madhukkarai – 641105, Coimbatore – Dist. on': {'and carried out survey'},
                                      'Corporate Office at Chennai – 600002, on': {'and subsequent dates'},
                                      'captioned loss was surveyed ': {'in the presence of'},
                                      'surveyed the next day i.e., on': {'. Damages'},
                                      'we surveyed the loss on': {'at above'},
                                      },

                   #                       'Excess':{'Less: Policy':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM'}
                   #                                  'POLICY EXCESS':{'ASSESSMENT OF LOSS','BACKGROUND'},
                   #                                  'Less: Policy':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss','Net Loss','Adjusted Amount'},
                   #                                  'Less : Policy':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss','Net Loss','Adjusted Amount'},
                   #                                  'LESS: POLICY':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','Net Adjusted','Net Adjusted Loss','NET ADIUSTED LOSS','Net Loss','Adjusted Amount'},
                   #                                  'LESS : POLICY':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','Net Adjusted','Net Adjusted Loss','NET ADIUSTED LOSS','Net Loss','Adjusted Amount'},
                   #                                  'Less:- Excess':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','Net Loss','Net Assessed Loss','NET ASSESSED LOSS','Net Adjusted','Net Adjusted Loss','NET ADJUSTED LOSS','NET LOSS','Adjusted Amount'},
                   #                                  'LESS: EXCESS':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','Net Loss','Net Assessed Loss','NET ASSESSED LOSS','Net Adjusted','Net Adjusted Loss','NET ADJUSTED LOSS','NET LOSS','Adjusted Amount'},
                   #                                  'LESS : EXCESS':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','Net Loss','Net Assessed Loss','NET ASSESSED LOSS','Net Adjusted','Net Adjusted Loss','NET ADJUSTED LOSS','NET LOSS','Adjusted Amount'},
                   #                                  'Less: Excess':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','Net Loss','Net Assessed Loss','NET ASSESSED LOSS','Net Adjusted Loss','NET ADJUSTED LOSS','for each','NET LOSS','Adjusted Amount'},
                   #                                  'Less : Excess':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','Net Loss','Net Assessed Loss','NET ASSESSED LOSS','Net Adjusted Loss','NET ADJUSTED LOSS','for each','NET LOSS','Adjusted Amount'},
                   #                                  'Less Policy':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','Net Assessed','Net Assessed Loss','NET ASSESSED LOSS','Net Assessed Loss','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss','Net Loss','Adjusted Amount'},
                   #                                  'LESS POLICY':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','Net Adjusted','Net Adjusted Loss','NET ADIUSTED LOSS','NET ADJUSTED LOSS','Net Loss','NET LOSS','Adjusted Amount'},
                   #                                  'LESS policy':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','Net Adjusted','Net Adjusted Loss','NET ADIUSTED LOSS','NET ADJUSTED LOSS','Net Loss','NET LOSS','Adjusted Amount'},
                   #                                  'Less: Policy':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss','Net Loss','Adjusted Amount','Net Adjusted Loss','Adjusted Amount','Net Payable Amount'}
                   #                                'Less: Policy':{'Net Assessed Loss','NET ASSESSED LOSS','NET ADJUSTED CLAIM'}
                   #                             },
                   'Excess': {
                       'Excess': {'Net loss payable'},
                       'claim amount subject to a minimum of': {'NET ADJUSTED CLAIM', 'Policy No',
                                                                'BACKGROUND & OCCURRENCE', 'Net Assessed Loss',
                                                                'CONSENT'},
                       'claim amount subject to minimum of': {'NET ADJUSTED CLAIM AMOUNT', 'Warranted'},
                       'Claim amount is subject to minimum of': {'NET ADJUSTED CLAIM'},
                       'claim amount subject to a minimum of Rs. 5.00 lakh': {'Net Assessed Loss'},
                       'Less: Excess @ 7 days Standard GP': {'Net Claim'},
                       'Million each and every loss': {'Net Assessed Loss'},
                       'subject to a minimum of ₹ 2500': {'Net Assessed Loss'},
                       'subject to a minimum of ₹ 25000': {'Net Assessed Loss'},
                       'claim amount subject to Min. Rs. 10,00,000': {'Net Assessed Loss'},
                       'claim amount subject to minimum of 10,000': {'Net Adjusted Amount'},
                       'amount subject to minimum of Rs.10 ': {'Net Assessed Loss'},
                       '7days Standard GP': {'Net Claim Payable'},
                       'subject to min. of Rs.': {'being deducted'},
                       'Less: Policy Excess @ 7 Days of Loss': {'Net Loss Payable'},
                       '8 Excess': {'9 Net Admissible [oss'},
                       'LESS: Policy Excess': {'Net Claim Payable'},
                       'Excess to be deducted is': {'..'},
                       'Policy Excess': {'Net LOP'},
                       'LESS EXCESS': {'CLAIM AMOUNT PAYABLE'},
                       'Less: Policy excess': {'Net Adjusted Loss', 'Regd. Office', 'Net Adiusted Loss',
                                               'Nei Adiusted Loss'},
                       'Less: Policy Excess': {'Adjusted Amount', 'IAR', 'Net Assessed Loss', 'Net Adjusted Loss',
                                               'Net Payable Amount'},
                       'Less policy excess': {'Net Assessed Loss'},
                       'Less Policy Excess': {'Net Assessed Loss', 'Net Adjusted Loss', 'Net Loss'},
                       'Less: Excess': {'Net Adjusted loss', '205025.39', 'Net Adjusted Claim Amount',
                                        'Net Adjusted Loss', 'Net Claim', 'Net Loss Assessed', 'Net Adjusted toss'},
                       'Less : Policy Excess': {'NET LOSS'},
                       'Less: llxcess': {'Net Adjusted Loss'},
                       'Less: Policv excess': {'Net Adiusted Loss'},
                       'Less Excess': {'Net Assessed Amount', 'Net Assessed Value', 'Net Adjusted Loss',
                                       'Net Adjusted Claim'},
                       'LESS: EXCESS': {'NETT LOSS'},
                       'Less:- Excess': {'Net Loss'},
                       'LESS: POLICY EXCESS': {'CONSIDERED', 'NET AD]USTED LOSS', 'NETADJUSTED LOSS',
                                               'NET ADIUSTED LOSS', 'NET ADJUSTED LOSS'},
                       'LESS: POLICY EXCESS': {'NET ADIUSTED LOSS', 'NETADJUSTED LOSS'},
                       'Less: Time Excess': {'Net Loss'},
                       'Less:Excess': {'Net Assessed Loss'},
                       'Less: MLOP P Excess': {'Net Loss'},
                       'LESS : Excess': {'Net Assessed Amount'}
                   },

                   'Final Amount': {'Nett Adjusted Loss': {'DESCRIPTION', 'the'}, 'Assessed Loss': {'INSURANCE'},
                                    'Adjusted Amount': {'INSURED’S CONSENT'}, 'NET ADJUSTED LOSS': 'This',
                                    'NET ADIUSTED LOSS': 'Survey & Loss',
                                    'Net Assessed Loss': {'INSURANCE', 'Page', 'NET ADJUSTED LOSS', 'The above',
                                                          'SUMMARY', 'This'}, 'Net Claim Amount': {'ANNEXURES'},
                                    'Net Claim Payable': {'CAUSE OF LOSS/DAMAGE'}},
                   'Gross Loss': {'GROSS LOSS ASSESSED': {'SALVAGE'},
                                  'Gross Loss': {'Less: Depreciation', 'Less ITC', 'Less:- Excess'},
                                  'Assessed Loss': {'INSURANCE'}, 'Assessed Loss': {'Less: Salvage'},
                                  'GROSS LOSS': {'Less : Credit', 'LE55: GST', 'LESS: UNDERINSURANCE', 'LESS: 10%',
                                                 'LESS: 1o9lo'}, 'LOSS ASSESSED (A+B)': 'Less'},
                   'Salvage': {'salvage value an amount of': {'damaged'},
                               '[ess: Salvage': 'Net Assessed loss', 'Less salvage LS basis': {'Value after Salvage'},
                               '5% i.e.': {'which is fair'},
                               'SALVAGE': {'UNDER INSURANCE', 'UNDERINSURANCE', 'EXCESS:'},
                               'Less: Salvage': {'Less: Under Insurance'},
                               'LESS: SALVAGE': {'NETASSESSED LOSS', 'NET ASSESSED LOSS'},
                               'Less Salvage': {'Assessed Loss', 'Assessed Loss netof'},
                               'Salvage Value': {'Net Assessed'}, 'SALVAGE VALUE': {'ASSESSMENT OF LOSS'},
                               'Less : Salvage': {'Less : Policy'},
                               'Less salvage LS basis': {'Value after'},
                               'Less: Value of salvage (Nil)': {'Gross Assessed', 'Gross Assessed Loss'},
                               'Less: Salvage Value': {'Sub Total', '7 Sub-Total'}},
                   #                'Excess':{'Less: Excess':{'NET ADJUSTED CLAIM','ADJUSTMENT','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss'},'Less: Policy excess':{'NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss'},'Less : Policy excess':{'NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss'},'Less : Policy Excess':{'NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET LOSS','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss'},'LESS: POLICY':{'NET ADJUSTED CLAIM','ADJUSTMENT','NET AD]USTED','NET ADJUSTED','NETADJUSTED LOSS','NET ADJUSTED LOSS','Net Adjusted','Net Adjusted Loss','NET ADIUSTED LOSS'},'Less:- Excess':{'NET ADJUSTED CLAIM','ADJUSTMENT','Net Loss','Net Adjusted','Net Adjusted Loss','NET ADJUSTED LOSS'},'Less: Excess':{'NET ADJUSTED CLAIM','ADJUSTMENT','Net Loss','Net Adjusted Loss','NET ADJUSTED LOSS','for each'},'Less Policy':{'NET ADJUSTED CLAIM','ADJUSTMENT','Net Assessed','NET ADJUSTED LOSS','NET LOSS','Net Adjusted Loss','Net Assessed Loss'},'Less policy':{'NET ADJUSTED CLAIM','ADJUSTMENT','Net Assessed','NET ADJUSTED LOSS','NET LOSS','Net Assessed Loss','NET ASSESSED LOSS','Net Adjusted Loss','NET ADJUSTED LOSS'}},
                   #                'Under Insurance':{'Underinsurance':'Offices:','UNDERINSURANCE':{'PROPORTIONATE LOSS','EXCESS'},'UNDER INSURANCE':{'EXCESS'},'Under Insurance Amount':{'Adjusted Loss'},'Less: Under Insurance':{'Sub Total','Adjusted Claim','Value after'},'Less Under Insurance':{'Sub Total','Adjusted Claim','Value after'},'LESS: UNDERINSURANCE':{'PROPORTIONATE LOSS'},'Less: UNDER INSURANCE':{'Sub Total','Adjusted Claim','Value after','PROPORTIONATE LOSS'},'LESS: UNDER INSURANCE':{'Sub Total','Adjusted Claim','Value after','PROPORTIONATE LOSS'}},
                   'Under Insurance': {'Underinsurance': {'Amount', 'Offices:'},
                                       'Less: Under Insurance': {'Adjusted Claim Amount', 'Regd.', 'Adjusted Loss',
                                                                 'Sub-Total'},
                                       'Less Underinsurance': {'Adjusted Loss'},
                                       'Less: Underinsurance': {'Adiusted Loss', 'Assessed Loss', 'Net Adjusted loss',
                                                                'Adjusted Loss'},
                                       'Under Insurance': {'Loss after'},
                                       'Under-lnsurance': {'6 uosi'},
                                       'LESS: Under Insurance': {'14 LESS: Policy Excess'},
                                       'LESS: UNDERINSURANCE': {'PROPORTIONATE LOSS'}

                                       },

                   #                'Depreciation':{'DEPRECIATION':'Offi6:','Less: Depreciation':{'Less Salvage','Assessed Loss','Loss Assessed'},'Less Depreciation':{'Assessed Loss','Loss Assessed','Less Salvage'},'Less: Depreciation':{'Assessed Loss','Loss Assessed','Less Salvage'},'LESS DEPRECIATION':{'Assessed Loss','Loss Assessed','Less Salvage'},'LESS: DEPRECIATION':{'Assessed Loss','Loss Assessed','Less Salvage'},'Depreciation In %':{'Loss Assessed','Less Salvage','Assessed Loss'}},
                   'Depreciation': {'Depreciation': {'Assessed Loss'},
                                    'Less: Depreciation': {'Less salvage', 'Assessed Loss', 'Value at Risk',
                                                           'Loss Assessed', '3 Less', 'Loss Assessed net'},
                                    'Lcss: Depreciation': {'Salvage Value'},
                                    'Less Depreciation': {'Loss Assessed'},
                                    'Obsolescence/Depreciation': {'Loss Assessed'},
                                    'Less: Obsolescence /Depreciation': {'Loss Assessed'},
                                    'Depreciation and amortization expense': {'Other expenses'},
                                    'Less : Depreciation': {'Less : Betterment'},
                                    },
                   'Debris Removal': {'Add: Removal of Debris': {'Adjusted Loss'}},
                   'Remarks': {'RECOMMNEDATION': {'The above rcpott'}, 'REMARKS:': {'ENCLOSURES:'}},
                   'Assessment (Full)': {
                       'NATURE, EXTENT OF': ['EARLIER RESERVE', 'DATE AND PLACE OF', 'CIRCUMSTANCES OF LOSS'],
                       'NATURE & APPARENT EXTENT OF DAMAGE': {'PHOTOGRAPHS'},
                       'NATURE & APPARENT': ['PHOTOGRAPHS', 'LIABILITY UNDER'],
                       'NATURE &': {'POLICY'},
                       'NATURE AND EXTENT OF DAMAGE': ['POLICY CONDITIONS/ WARRANTIES', 'POLICY LIABILITY'],
                       'EXTENT OF STOLEN/DAMAGE': {'OUR OBSERVATION'},
                       'BACKGROUND & OCCURRENCE:': {'OUR SURVEY'},
                       'INSPECTION': {'OPINION ON'},
                       'Nature and Extent of': 'Cause of'},
                   'Property Damaged': {'# Description': {'As per policy'}},
                   #                 'Remarks':{'RECOMMENDATION':{'Mumbai Office'}}
                   #             'Amount per Quantity':{'SL DESCRIPTION QTY AMOUNT (RS)':{'GROSS ASSESSED LOSS'}}
                   #             'Property Damaged_ASS':{'ASSESSMENT OF LOSS:-':'11.00 SALVAGE:-'},
                   #             'Quantity_ASS':{'ASSESSMENT OF LOSS':'SALVAGE:-'}
                   #             'Unit_ASS':{'':''}
                   #             'Amount per Quantity_ASS':{'':''}
                   'Total Amount_ASS': {'ASSESSABLE LOSS AMAUNT (INR) ': {'IAR-1804-16700_Reliance'}},
                   #             'Property Damaged':{'INSURED’S CLAIM:-':'ASSESSMENT OF LOSS'},
                   #             'Quantity':{'':''}
                   #             'Unit':{'':''}
                   #             'Amount per Quantity':{'':''}
                   'Total Amount': {'Total Amount A + B': {'Adjustment Sheet'}}

                   }
        out = {'IL Claim Ref No': None, 'Date of Loss': None, 'Cause of Loss': None,
               'Cause of Loss (Sentence) Description': None, 'Estimate of Loss': None, 'Loss Location (City)': None,
               'Loss Location (Full) Address': None, 'Date of Survey': None, 'Property Damaged_ASS': None,
               'Quantity_ASS': None, 'Unit_ASS': None, 'Amount per Quantity_ASS': None, 'Total Amount_ASS': None,
               'Property Damaged': None, 'Quantity': None, 'Unit': None, 'Amount per Quantity': None,
               'Total Amount': None, 'Gross Loss': None, 'Depreciation': None, 'Excess': None, 'Salvage': None,
               'Debris Removal': None, 'Under Insurance': None, 'Final Amount': None, 'Assessment (Full)': None,
               'Remarks': None}
        # out={}
        exu = 0
        for k1 in dic_key:
            exu = 0
            for k2 in dic_key[k1]:
                vs = []
                for x in dic_key[k1][k2]:

                    if x and k2 in txt:
                        clm_no = ((txt.split(k2)[1]).split(x)[0]).strip()
                        # print('********************',k1,k2,x)27/28-08-2020
                        if k1 == 'Date of Loss':
                            clm_no = clm_no.replace('\'', '')
                            temp = re.findall(
                                "\d{2}\s\w{3,8}\s\d{4}|\d{2}/?.?-?\d{2}/?.?-?\d{4}|\d{2}\w{2}\s\w{3,9}\s\d{4}|\d{2}/\d{2}/?.?-?\d{2}/?.?-?\d{4}|\d{2}/?.?-?\d{2}/?.?-?\d{2}|\d{2}-\w{3}-\d{2}",
                                clm_no)
                            if len(temp) >= 1 and len(temp[0]) > 8:
                                clm_no = temp[0]
                                temp = clm_no
                                if 'Date' in clm_no:
                                    clm_no = clm_no.split('Date')
                                if (temp.isdigit() or (re.sub(' ', '', temp)).isdigit() or temp.count(
                                        '.') == 1 or temp.count('/') == 1 or temp.count('-') == 1):
                                    clm_no = ''
                            else:
                                clm_no = ''

                        if k1 == 'IL Claim Ref No':
                            if k2 == 'Green house claims':
                                temp = re.findall("\w{2,4}\d{7,10}", clm_no)
                                clm_no = temp[0]
                            elif k2 == 'CLAIM REFERENCE':
                                clm_no = clm_no.strip()
                            else:
                                temp = re.findall("\w{2,4}\d{7,10}", clm_no)
                                if len(temp) >= 1:
                                    clm_no = temp[0].replace('\'', '')
                                    temp = clm_no
                                    if temp.isdigit():
                                        clm_no = ''
                                else:
                                    clm_no = ''

                        if k1 == 'Cause of Loss':
                            if 'Point of origin' in clm_no or 'How the loss' in clm_no:
                                clm_no = ((txt.split('CIRCUMSTANCES')[1]).split('ESTIMATED AMOUNT')[0]).strip()

                            else:
                                clm_no = clm_no.replace(':', '').replace('OF LOSS', '').replace('LOR', '').replace(
                                    'LOSS', '').strip()
                                if 'EARLIER' in clm_no:
                                    clm_no = clm_no.split('EARLIER')[0]
                                if 'GROSS ASSESSED' in clm_no:
                                    clm_no = clm_no.split('GROSS ASSESSED')[0]
                                if 'LOSS ASSESSMENT' in clm_no:
                                    clm_no = clm_no.split('LOSS ASSESSMENT')[0]
                                if 'ASSESSED' in clm_no:
                                    clm_no = clm_no.split('ASSESSED')[0]
                                if 'SUBROGATION' in clm_no:
                                    clm_no = clm_no.split('SUBROGATION')[0]
                                if 'NATURE' in clm_no:
                                    clm_no = clm_no.split('NATURE')[0]
                                if '01 of 06' in clm_no:
                                    clm_no = clm_no.split('01 of 06')[0]
                                if 'CIRCUMSTANCES' in clm_no:
                                    clm_no = clm_no.split('CIRCUMSTANCES')[0]
                                if 'REPORT' in clm_no:
                                    clm_no = clm_no.split('REPORT')[0]
                                if 'OCCUPANCY' in clm_no:
                                    clm_no = clm_no.split('OCCUPANCY')[0]
                                if '01 of 05' in clm_no:
                                    clm_no = clm_no.split('01 of 05')[0]
                                if 'ADDRESS' in clm_no:
                                    clm_no = clm_no.split('ADDRESS')[0]
                                if 'Reg' in clm_no:
                                    clm_no = clm_no.split('Reg')[0]
                                if 'DATE' in clm_no:
                                    clm_no = clm_no.split('DATE')[0]
                                if 'PROPERTY' in clm_no:
                                    clm_no = clm_no.split('PROPERTY')[0]
                                if ', AS' in clm_no and len(clm_no.split(', AS')) >= 2:
                                    clm_no = clm_no.split(', AS')[1]

                        if k1 == 'Cause of Loss (Sentence) Description':
                            clm_no = ''
                            if 'CAUSE OF LOSS' and 'ASSESSMENT' in txt and len(txt.split('CAUSE OF LOSS')) >= 2:
                                clm_no = ((txt.split('CAUSE OF LOSS')[1]).split('ASSESSMENT')[0]).strip()
                            if 'CIRCUMSTANCES OF INCIDENT' in txt and 'Cause of loss' not in txt and len(
                                    txt.split('CIRCUMSTANCES OF INCIDENT')) >= 2 and 'Page' in txt:
                                clm_no = ((txt.split('CIRCUMSTANCES OF INCIDENT')[1]).split('Page')[0]).strip()
                            if 'CIRCUMSTANCES/DISCOVERY' in txt and 'Cause of loss' not in txt and len(
                                    txt.split('CIRCUMSTANCES/DISCOVERY')) >= 2:
                                clm_no = ((txt.split('CIRCUMSTANCES/DISCOVERY')[1]).split('OCCURENCE')[0]).strip()
                            if 'Cause of loss' in txt and len(txt.split('Cause of loss')) >= 2:
                                clm_no = ((txt.split('Cause of loss')[1]).split('PHOTOGRAPHS OF')[0]).strip()
                            clm_no = clm_no.replace('LOSS', '')
                            clm_no = clm_no.replace(':', '')

                        #                             if 'OF LOSS' in clm_no:
                        #                                 clm_no=((txt.split('NATURE & EXTENT OF DAMAGE')[1]).split('POLICY CONDITIONS')[0]).strip()

                        #                             if 'IAR' and 'of' in clm_no:
                        #                                 temp=clm_no[clm_no.find(' IAR'):clm_no.find('of ')+5]
                        #                                 clm_no=clm_no.replace(temp,'').strip()
                        #                             clm_no=clm_no.rsplit('(',1)[0].strip().replace(':','')
                        #                             if 'OUR SURVEY' in clm_no:
                        #                                 clm_no=clm_no.split('OUR SURVEY')[0]
                        #                             if 'Page' in clm_no:
                        #                                 temp=clm_no[clm_no.find('Page'):clm_no.find('Page')+9]
                        #                                 clm_no=clm_no.replace(temp,'')

                        if k1 == 'Estimate of Loss':
                            clm_no = clm_no.replace('\'', '')
                            temp = re.findall("\d+,?\d+.?\d+", clm_no)
                            if len(temp) > 0 and ('NIL' in clm_no or 'Nil' in clm_no):
                                if clm_no.find(temp[0]) > (clm_no.find('NIL') or clm_no.find('Nil')):
                                    clm_no = 'NIL'
                            if len(temp) > 0:
                                if '%' not in temp[0]:
                                    clm_no = temp[0].split(' ')[0]
                                elif len(temp) >= 2 and '%' not in temp[1]:
                                    clm_no = temp[1].split(' ')[0]
                                temp = clm_no
                                if (temp.replace(',', '').replace('.', '')).isdigit() and clm_no.count('.') <= 1:
                                    clm_no = temp
                                else:
                                    clm_no = ''
                            else:
                                clm_no = ''

                        if k1 == 'Loss Location (Full) Address':
                            clm_no = clm_no.replace('Type', '').replace(') M/s', '').replace('LOCATION, WITH PIN',
                                                                                             '').replace('CODE',
                                                                                                         '').replace(
                                'UNDER POLICY', '').strip()
                            if 'Sq.m at' in clm_no:
                                clm_no = (clm_no.split('Sq.m at')[1]).strip()
                            if '2019 at' in clm_no and len(clm_no.split('2019 at')) >= 2:
                                clm_no = clm_no.split('2019 at')[1]
                            if '2020 at' in clm_no and len(clm_no.split('2020 at')) >= 2:
                                clm_no = clm_no.split('2020 at')[1]
                            if '2020at' in clm_no and len(clm_no.split('2020at')) >= 2:
                                clm_no = clm_no.split('2020at')[1]

                            if 'DETAILS' in clm_no:
                                clm_no = (clm_no.split('DETAILS')[0]).strip()
                            if 'Occupancy' in clm_no:
                                clm_no = (clm_no.split('Occupancy')[0]).strip()
                            if 'PERSON' in clm_no:
                                clm_no = clm_no.split('PERSON')[0]
                            if 'CAUSE' in clm_no:
                                clm_no = clm_no.split('CAUSE')[0]
                            if 'DATE' in clm_no:
                                clm_no = clm_no.split('DATE')[0]
                            if 'Page' in clm_no:
                                clm_no = clm_no.split('Page')[0]
                            if 'NATURE' in clm_no:
                                clm_no = clm_no.split('NATURE')[0]
                            if 'DESCRIPTION' in clm_no:
                                clm_no = clm_no.split('DESCRIPTION')[0]
                            if 'DETAILS' in clm_no:
                                clm_no = clm_no.split('DETAILS')[0]
                            if 'Description' in clm_no:
                                clm_no = clm_no.split('Description')[0]

                            if '2. POLICY' in clm_no:
                                clm_no = clm_no.split('2. POLICY')[0]
                            if 'OCCUPANCY' in clm_no:
                                clm_no = clm_no.split('2. POLICY')[0]

                        if k1 == 'Date of Survey':
                            clm_no = clm_no.replace('\'', '')
                            temp = re.findall(
                                "\d{2}\s\w{3,8}\s\d{4}|\d{2}/?.?-?\d{2}/?.?-?\d{4}|\d{2}\w{2}\s\w{3,9}\s\d{4}|\d{2}/\d{2}/?.?-?\d{2}/?.?-?\d{4}|\d{2}-\w{3}-\d{2}",
                                clm_no)
                            if len(temp) >= 1 and len(temp[0]) > 8:
                                clm_no = temp[0]
                                temp = clm_no
                                if temp.isdigit() or (re.sub(' ', '', temp)).isdigit() or temp.count(
                                        '.') == 1 or temp.count('/') == 1 or temp.count('-') == 1:
                                    clm_no = ''
                            else:
                                clm_no = ''

                        if k1 == 'Final Amount':
                            clm_no = clm_no.replace('\'', '')
                            if '₹' in clm_no:
                                clm_no = ''
                                clm_no = ((txt.split('Net Assessed Loss')[1]).split('Annexure')[0]).strip()
                                clm_no = clm_no.strip().split('₹')[1].lstrip().split(' ')[0].replace('/', '').replace(
                                    '-', '').replace('.', '')
                            temp = re.findall("\d+,?.*?\d+.?\d+", clm_no)
                            if len(temp) > 0 and ('NIL' in clm_no or 'Nil' in clm_no):
                                if clm_no.find(temp[0]) > (clm_no.find('NIL') or clm_no.find('Nil')):
                                    clm_no = 'NIL'
                            if len(temp) > 0:
                                if '%' not in temp[0]:
                                    clm_no = temp[0].split(' ')[0]
                                elif len(temp) >= 2 and '%' not in temp[1]:
                                    clm_no = temp[1].split(' ')[0]
                                temp = clm_no
                                if (temp.replace(',', '').replace('.', '')).isdigit() and clm_no.count('.') <= 1:
                                    clm_no = temp
                                else:
                                    clm_no = ''
                            else:
                                clm_no = ''

                        if k1 == 'Gross Loss':
                            clm_no = clm_no.replace('\'', '')
                            temp = re.findall("\d+,?.*?\d+.?\d+", clm_no)
                            if len(temp) > 0 and ('NIL' in clm_no or 'Nil' in clm_no):
                                if clm_no.find(temp[0]) > (clm_no.find('NIL') or clm_no.find('Nil')):
                                    clm_no = 'NIL'
                            if len(temp) > 0:
                                if '%' not in temp[0]:
                                    clm_no = temp[0].split(' ')[0]
                                else:
                                    if len(temp) >= 2 and '%' not in temp[1]:
                                        clm_no = temp[1].split(' ')[0]
                                temp = clm_no
                                if (temp.replace(',', '').replace('.', '')).isdigit() and clm_no.count('.') <= 1:
                                    clm_no = temp
                                else:
                                    clm_no = ''
                            else:
                                clm_no = ''

                        if k1 == 'Salvage':
                            clm_no = clm_no.replace('\'', '')
                            temp = re.findall("\d+,?.*?\d+.?\d+", clm_no)
                            if len(temp) > 0 and ('NIL' in clm_no or 'Nil' in clm_no):
                                if clm_no.find(temp[0]) > (clm_no.find('NIL') or clm_no.find('Nil')):
                                    clm_no = 'NIL'
                            if len(temp) > 0:
                                if '%' not in temp[0]:
                                    clm_no = temp[0].split(' ')[0]
                                else:
                                    if len(temp) >= 2 and '%' not in temp[1]:
                                        clm_no = temp[1].split(' ')[0]
                                temp = clm_no
                                if (temp.replace(',', '').replace('.', '')).isdigit() and clm_no.count('.') <= 1:
                                    clm_no = temp
                                else:
                                    clm_no = ''

                            else:
                                clm_no = ''

                        if k1 == 'Under Insurance':
                            clm_no = clm_no.replace('\'', '')
                            temp = re.findall("\d+,?.*?\d+.?\d+", clm_no)
                            if len(temp) > 0 and ('NIL' in clm_no or 'Nil' in clm_no):
                                if clm_no.find(temp[0]) > (clm_no.find('NIL') or clm_no.find('Nil')):
                                    clm_no = 'NIL'
                            if len(temp) > 0:
                                if '%' not in temp[0] and len(temp) == 1 and len(temp[0]) > 3:
                                    clm_no = temp[0].split(' ')[0]
                                else:
                                    if len(temp) >= 2 and '%' not in temp[1] and len(temp[1]) > 3:
                                        clm_no = temp[1].split(' ')[0]
                                temp = clm_no
                                if (temp.replace(',', '').replace('.', '')).isdigit() and clm_no.count('.') <= 1:
                                    clm_no = temp
                                else:
                                    clm_no = ''
                            else:
                                clm_no = ''

                        if k1 == 'Excess':
                            #                             clm_no=clm_no.replace('\'','')

                            temp = re.findall("\d+,?\d+.?\d+", clm_no)
                            if len(temp) > 0 and ('NIL' in clm_no or 'Nil' in clm_no):
                                if clm_no.find(temp[0]) > (clm_no.find('NIL') or clm_no.find('Nil')):
                                    clm_no = 'NIL'
                            if len(temp) > 0:
                                if '%' not in temp[0]:
                                    clm_no = temp[0].split(' ')[0]
                                #                                 else:
                                #                                     if len(temp)>=2 and '%' not in temp[2]:
                                #                                         clm_no=temp[2].split(' ')[0]
                                temp = clm_no
                                if (temp.replace(',', '').replace('.', '')).isdigit() and clm_no.count(
                                        '.') <= 1 and '-' not in clm_no:
                                    clm_no = temp
                                else:
                                    clm_no = ''
                            else:
                                clm_no = ''
                        if k1 == 'Depreciation':
                            clm_no = clm_no.replace('\'', '')
                            temp = re.findall("\d+,?.*?\d+.?\d+", clm_no)
                            if len(temp) > 0 and ('NIL' in clm_no or 'Nil' in clm_no):
                                if clm_no.find(temp[0]) > (clm_no.find('NIL') or clm_no.find('Nil')):
                                    clm_no = 'NIL'
                            if len(temp) > 0:
                                if '%' not in temp[0] and len(temp[0]) > 3:
                                    clm_no = temp[0].split(' ')[0]
                                else:
                                    if len(temp) >= 2 and '%' not in temp[1] and len(temp[1]) > 3:
                                        clm_no = temp[1].split(' ')[0]
                                temp = clm_no
                                if (temp.replace(',', '').replace('.', '')).isdigit() and clm_no.count('.') <= 1:
                                    clm_no = temp
                                else:
                                    clm_no = ''
                            else:
                                clm_no = ''

                        if k1 == 'Total Amount_ASS':
                            clm_no = clm_no.strip()
                        if k1 == 'Total Amount':
                            clm_no = clm_no.strip()
                        if k1 == 'Remarks':
                            clm_no = clm_no.strip()
                        if k1 == 'Debris Removal':
                            clm_no = clm_no.strip()

                        if k1 == 'Assessment (Full)':

                            if 'NATURE &EXTENT OF DAMAGE' in txt and len(txt.split('NATURE &EXTENT OF DAMAGE')) >= 2:
                                clm_no = ''
                                clm_no = txt.split('NATURE &EXTENT OF DAMAGE')[1]
                                if 'CAUSE OF LOSS' in clm_no:
                                    clm_no = txt.split('CAUSE OF LOSS')[0]
                            if 'NATURE & EXTENT OF DAMAGE' in txt and len(txt.split('NATURE &EXTENT OF DAMAGE')) >= 2:
                                clm_no = ''
                                clm_no = txt.split('NATURE &EXTENT OF DAMAGE')[1]
                                if 'CAUSE OF LOSS' in clm_no:
                                    clm_no = txt.split('CAUSE OF LOSS')[0]
                            if 'NATURE & EXTENT OF DAMAGE' in txt and len(txt.split('NATURE & EXTENT OF DAMAGE')) >= 2:
                                clm_no = ''
                                clm_no = txt.split('NATURE & EXTENT OF DAMAGE')[1]
                                if 'POLICY CONDITIONS' in clm_no:
                                    clm_no = txt.split('POLICY CONDITIONS')[0]
                            clm_no = clm_no.replace(':', '').strip()

                        #                             if 'IAR' and 'of' in clm_no:
                        #                                 temp=clm_no[clm_no.find(' IAR'):clm_no.find('of ')+5]
                        #                                 clm_no=clm_no.replace(temp,'').strip()
                        #                             if '6.00' in clm_no:
                        #                                 clm_no=clm_no.replace('6.00','').strip()
                        #                         clm_no=clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        #                         if 'Date' in clm_no:
                        #                             clm_no=''
                        #                             clm_no=txt.split('Cause of Loss')[0]
                        #                             if 'Cash Loss' not in clm_no:
                        #                                 clm_no=(clm_no.rsplit('Loss',1)[1]).strip()
                        #                                 clm_no=clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        #                             elif 'Cash Loss' in clm_no:
                        #                                 clm_no='Cash Loss'
                        #                         clm_no=clm_no.replace('IAR-1805-17396_Reliance Retail Ltd_FSR','').replace('Page 3 of 4','').strip()
                        if k1 == 'Property Damaged_ASS':
                            clm_no = clm_no.strip()

                        if k1 == 'Property Damaged':
                            clm_no = clm_no.replace('IAR-1805-17396_Reliance Retail Ltd_FSR', '').replace('Page 3 of 4',
                                                                                                          '').strip()
                        #                         clm_no=clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        #                         if 'Date' in clm_no:
                        #                             clm_no=''
                        #                             clm_no=txt.split('Cause of Loss')[0]
                        #                             if 'Cash Loss' not in clm_no:
                        #                                 clm_no=(clm_no.rsplit('Loss',1)[1]).strip()
                        #                                 clm_no=clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        #                             elif 'Cash Loss' in clm_no:
                        #                                 clm_no='Cash Loss'
                        #                         c_new=[]
                        #                         c_new_item=[]
                        #                         c_new_amt=[]
                        #                         for r in re.split('\d+\s',clm_no):
                        #                             if len(r)==0:
                        #                                 continue
                        #                             c_new.append(r)
                        #                         for r in range(len(c_new)):
                        #                             if r%2==0:
                        #                                 c_new_item.append(c_new[r])
                        #                             else:
                        #                                 c_new_amt.append(c_new[r])
                        #                         item=''
                        #                         amt=''
                        #                         for c in c_new_item:
                        #                             item +=c +', '
                        #                         item=item.strip(', ')
                        #                         for c in c_new_amt:
                        #                             amt +=c +', '
                        #                         amt=amt.strip(', ')
                        #                         clm_no=item
                        if k1 == 'Amount per Quantity':
                            c_new = []
                            c_new_item = []
                            c_new_amt = []
                            for r in re.split('\d+\s', clm_no):
                                if len(r) == 0:
                                    continue
                                c_new.append(r)
                            for r in range(len(c_new)):
                                if r % 2 == 0:
                                    c_new_item.append(c_new[r])
                                else:
                                    c_new_amt.append(c_new[r])
                            item = ''
                            amt = ''
                            for c in c_new_item:
                                item += c + ', '
                            item = item.strip(', ')
                            for c in c_new_amt:
                                amt += c + ', '
                            amt = amt.strip(', ')
                            clm_no = amt
                        if len(vs) == 0:
                            vs.append(len(clm_no))
                            vs.append(clm_no)
                            out[k1] = vs[1]
                            exu = 1
                            continue
                        if len(vs) != 0:
                            if len(clm_no) > vs[0]:
                                continue
                            vs = []
                            vs.append(len(clm_no))
                            vs.append(clm_no)
                            out[k1] = vs[1]
                            exu = 1

                if exu == 1:
                    break

        print(fl, '#########################################################################')
        city = ''
        if out['Loss Location (Full) Address'] is not None:
            city = city_finder(out['Loss Location (Full) Address'])
            out['Loss Location (City)'] = city

        doc_num = ''
        new_link = ''
        if out['IL Claim Ref No'] is not None:
            doc_num = out['IL Claim Ref No']
        if len(doc_num) > 5:
            doc_num = re.sub('[^A-Z0-9]', '', doc_num)
            # doc_num=
            print(doc_num, 'lllll')
        for d in data:
            if doc_num == d.split('&$')[0]:
                print('yyyyyyyyyy', d.split('&$')[2])
                new_link = d.split('&$')[2]
                # new_link='new_link'
                break
        for d in data_one:
            if doc_num == d.split('&$')[0]:
                # out['Estimate of Loss']="No Claim"
                out['Gross Loss'] = "Nil"
                out['Salvage'] = "Nil"
                out['Excess'] = "Nil"
                out['Final Amount'] = "Nil"
                out['Under Insurance'] = "Nil"
                out['Depreciation'] = "Nil"
                out['Debris Removal'] = "Nil"

        print(out)
        mbk = openpyxl.load_workbook(path, read_only=False)
        sht = mbk.get_sheet_by_name('Sheet1')
        p = 3 + cnt
        cnt += 1
        sht.cell(row=p, column=1).value = "IAR"
        sht.cell(row=p, column=2).hyperlink = fl
        sht.cell(row=p, column=2).value = fl.split('.txt')[0]
        c = 3
        for t in out:
            sht.cell(row=p, column=c).value = out[t]
            c += 1
        mbk.save(path)