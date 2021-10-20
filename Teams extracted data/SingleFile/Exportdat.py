import mysql.connector
from datetime import date, datetime

from colorama import Style
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Side
from openpyxl.styles import Alignment
from xlrd import sheet

def connection(today):
    conn = mysql.connector.connect(host='localhost', user='root', password='qazwsx@12340', database='starhealth_v1_db')
    myconnection = conn.cursor()
    myconnection.execute(
        "select count(*) from xxsh_uw_proposal_hdr where status='Awaiting Underwriter Approval'and creation_date={}".format(
            today))
    rowAwait = myconnection.fetchall()
    myconnection.execute(
        "select count(*) from xxsh_uw_proposal_hdr where status='Rejected'and creation_date={}".format(today))
    rowsRejected = myconnection.fetchall()
    myconnection.execute(
        "select count(*) from xxsh_uw_proposal_hdr where status='Success'and creation_date={}".format(today))
    rowsSuccess = myconnection.fetchall()
    myconnection.execute(
        "select count(*) from xxsh_uw_proposal_hdr where status='Pending'and creation_date={}".format(today))
    rowsPending = myconnection.fetchall()
    myconnection.execute(
        "select count(*) from xxsh_uw_proposal_hdr where status='Processing'and creation_date={}".format(today))
    rowsProcessing = myconnection.fetchall()

    return rowAwait[0][0],rowsRejected[0][0],rowsSuccess[0][0],rowsPending[0][0],rowsProcessing[0][0],rowAwait[0][0],rowsProcessing[0][0],rowAwait[0][0],rowAwait[0][0],rowAwait[0][0],rowAwait[0][0],rowAwait[0][0],rowAwait[0][0]


def mergecells(w1):
    w1.merge_cells('A1:A4')
    w1.merge_cells('A11:A13')
    w1.merge_cells('E2:F2')
    w1.merge_cells('E11:F11')
    w1.merge_cells('B2:D2')
    w1.merge_cells('G2:H2')
    w1.merge_cells('K3:K18')


SFill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('000066'))
SAlignment = Alignment(horizontal='center', vertical='center',wrapText=True)
SFont = Font(color="FFFFFF",size=12)
thick_border = Border(left=Side(border_style='dashed',color='FFFFFF'),
                right=Side(border_style='dashed',color='FFFFFF'),
                top=Side(border_style='thin',color='FFFFFF'),
                bottom=Side(border_style='thin',color='FFFFFF')
                )
Double_border = Border(left=Side(border_style='double',color='000000'),
                right=Side(border_style='double',color='000000'),
                top=Side(border_style='double',color='000000'),
                bottom=Side(border_style='double',color='000000')
                )
wb = Workbook()
def SummaryXldata(rowAwait, rowsRejected, rowsSuccess, rowsPending, rowsProcessing,rowsUnderwriter,rowsAgent,rowSenttoVO,rowsSuccessful,rowsFailed,rowsSuccess2,rowsFailure,rowsReporcessed):
    lista = [rowAwait, rowsRejected, rowsSuccess, rowsPending, rowsProcessing,rowsUnderwriter,rowsAgent]
    listb = [rowSenttoVO,rowsSuccessful,rowsFailed,rowsSuccess2,rowsFailure,rowsReporcessed]
    ListofINFLOW = ['Received', 'Approved', 'Rejected', 'Inaction', 'Missing Documents',
                    'Underwriter (24Hrs)', 'Agent (48 Hrs)']
    ListofOUTFLOW = ['Sent to VO', 'Successful', 'Failed', 'Success', 'Failure', 'Reporcessed']
    w1 = wb.active
    w1.title = "Summary.xlsx"
    for row in range(len(lista)):
        _ = w1.cell(column=row + 2, row=4, value=lista[row])
        _.alignment = SAlignment
        _.font = Font(color="000066")
    for row in range(len(listb)):
        _ = w1.cell(column=row + 2, row=13, value=listb[row])
        _.alignment = SAlignment
        _.font = Font(color="000066")
    for row in range(len(ListofINFLOW)):
        _ = w1.cell(column=row + 2, row=3, value=ListofINFLOW[row])
        _.alignment = SAlignment
        _.fill = SFill
        _.font = SFont
        _.border = thick_border
        _ = w1.cell(column=row + 2, row=2)
        _.fill = SFill
        _.alignment = SAlignment
        _.font = SFont
        _.border = thick_border
        _ = w1.cell(column=row + 2, row=1)
        _.fill = SFill

    for row in range(len(ListofOUTFLOW)):
        _ = w1.cell(column=row + 2, row=12, value=ListofOUTFLOW[row])
        _.alignment = SAlignment
        _.fill = SFill
        _.font = SFont
        _.border = thick_border
        _ = w1.cell(column=row+2, row=11)
        _.fill = SFill
        _.font = SFont
        _.border = thick_border
    w1["A1"].value = 'INFLOW'
    w1["A1"].alignment = SAlignment
    w1["B2"].value = 'Proposals by Status'
    w1["E2"].value = 'Proposal Pending'
    w1["G2"].value = 'Missed SLA'
    w1["A11"].value = 'OUTFLOW'
    w1["A11"].alignment = SAlignment
    w1["E11"].value = 'Average time'
    w1["E11"].alignment = SAlignment
    w1["K3"].value = 'Drill Through'
    w1["k3"].fill=SFill
    w1["K3"].font=SFont
    w1["K3"].alignment = SAlignment
    mergecells(w1)

def RPAdata():
    ListOfHeader=['Reference Number','Proposal Type','Status','Start time','End time','Missing docs','Source','If Failed Reason']
    w2=wb.create_sheet(title='RPA.xlsx')
    for row in range(len(ListOfHeader)):
        _ = w2.cell(column=row + 2, row=5, value=ListOfHeader[row])
        _.alignment = SAlignment
        _.border =Double_border


def main_function():
    excel_filename = datetime.now().strftime("%Y-%m-%d") + ".xlsx"
    today = date.today()
    rowAwait, rowsRejected, rowsSuccess, rowsPending, rowsProcessing,rowsUnderwriter,rowsAgent,rowSenttoVO,rowsSuccessful,rowsFailed,rowsSuccess2,rowsFailure,rowsReporcessed = connection(today)
    SummaryXldata(rowAwait, rowsRejected, rowsSuccess, rowsPending, rowsProcessing,rowsUnderwriter,rowsAgent,rowSenttoVO,rowsSuccessful,rowsFailed,rowsSuccess2,rowsFailure,rowsReporcessed)
    RPAdata()
    wb.save(excel_filename)


if __name__ == '__main__':
    main_function()