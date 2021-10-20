import PyPDF2
import os
import re
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import openpyxl
import csv


# files = [each for each in os.listdir(r'C:\Users\DEBJYOTI BANERJEE\Downloads\TEAM Final\TEAM') if each.endswith('.txt')]
files = [each for each in os.listdir(r'C:\Teams extracted data\SingleFile') if each.endswith('.txt')]
path = 'PROTOCOL.xlsx'
cnt = 0

# ///before loop starts
filename1 = "Excel_Files.csv"
with open(filename1, newline='', encoding='utf-8', errors='ignore') as f:
    reader = csv.reader(f)
    data1 = list(reader)
data = []
cn=0
for elem in data1[1:    ]:
    string1 = ''
    for item2 in elem:
        string1 = string1 + item2 + '&$'
    data.append(string1)
# print(data)
filename2 = "NIL Claim.csv"
with open(filename2, newline='', encoding='utf-8', errors='ignore') as f:
    reader = csv.reader(f)
    data1 = list(reader)
data_one = []
cn=0
for elem in data1[1:]:
    string1 = ''
    for item2 in elem:
        string1 = string1 + item2 + '&$'
    data_one.append(string1)
# print(data_one)


for fl in files:
    txt = ''
    with open(fl, mode='rt', encoding='utf-8') as f:
        for l in f:

            txt += l + ' '

    txt = re.sub(' +', ' ', txt)
    txt = txt.replace('\n', '')
    txt = txt.replace('\f', '')



    dic_key = {'IL Claim Ref No': {'Claim No.:':[')','Insured','Ref'],'Underwriters Ref No.:':['Policy no'],'Claim Ref No.':['Ref. No'],'Chim No.': ['ICICI'],"Insurer’s Ref. No.:":['MATERIAL'] ,"Insurer's Claim No.": ['ICICI Lombard', 'M/s'], 'Claim No.': ['Insured','Ref.',')','IMMEDIATE LOSS', 'ISSUED WITHOUT', 'Insured Claim no.',
                                                                                           'ICICI Lombard', 'FINAL SURVEY REPORT', 'Utsav Apartments','Impact Loss',
                                                                                           'INTERIM REPORT', '-', 'Period Of', 'WITHOUT PREJUDICE', 'Date', 'Breakdown', 'Submitted to',
                                                                                                  'Issued Without', 'Policy', 'THE INSURED', 'M/s', 'Claimed Considered', 'lClCl', ', Team Ref', 'Subject Matter', '1.0 Pursuant', 'ompany Ltd', 'Interim Report'],
                                   'CLAIM NO.': ['PREVIOUS REPORT', 'NAME & ADDRESS OF', 'POLICY TYPE', 'POLICY NO', 'CAUSE OF LOSS', 'INSURED', 'SURVEYOR APPOINTMENT', 'REASON FOR', '(Without', '3.00 INCIDENT', 'Dear Sirs', 'NAME OF', 'Ref. No', 'DATE OF LOSS'], 'CLAIM NO': ['POLICY TYPE', 'Regd. Office', 'RCS REF', ']'],
                                   'INSURER’S CLAIM NO.': ['POLICY TYPE/NO.'], 'Claim no.': ['lClCl', 'ICICI'],
                                   'ChiD oo.': ['ICICI Lombrrd'], 'INSURER’S CLAIM REF NO.': ['POLICY NO'], 'Claim RefNo. :-': ['POLICY'],
                                   'Claim Ref No. :-': ['[External Email', 'Dear', ', Team ref', ', Our Ref']},
               'Date of Loss': {'Furnace on ':['at M/s'],'Shahajan':[' U.P.)'],'reported, on':['at 00:40 A.M'],'Monday,':['Plot no'],'circle dated':['.'],'Fire Loss dated': ['under Standard'], 'Loss date': ['No of days', 'Item'], 'Loss Date': ['Item', 'Cause Of Loss'],  'TIME, DAY & DATE OF': {'INCIDENT', 'as reported'},
                                'Date of Loss': ['at around','Peril Reported','Banker/', 'Date of Intimation', 'at around', 'Date of', 'Financer', 'Insured'], 'DATE OF INCIDENT': {'LOSS INTIMATION'},
                                'TIME AND DATE OF LOSS': ['CLAIM INTIMATION DETAILS'], 'Time &Date of On': ['at about'],'DATE & TIME OF LOSS': ['DATE OF INTIMATION', 'Regd. Office', 'at morning hours', 'DATE OF SURVEY'],
                                "DATE OF LOSS": ['CAUSE OF LOSS', 'DATE OF INTIMATION', 'CLAIM NO.', 'DATE OF INTIMATIoN', 'DETAILS OF', 'DATE OF SURVEY', '3.02 NATURE'],
                                'inundation / flooding on': [', reported to have occurred'], 'date of loss': ['. The particular'], 'TIME& DATE OF': ['INCIDENT'], 'Date of loss': [', Claim No'],
                                'Date & Time of Loss': ['Date of Intimation']},
               'Cause of Loss': {'shut down on':['an 8 at'],'CAUSE OF LOSS AND LIABILITY UNDER THE POLICY': ['In this regard'], '6.2 CAUSE OF LOSS': ['6.2.2 As per'], 'CAUSE OF LOSS': ['10.00','NATURE, EXTENT OF', 'NATURE & EXTENT', 'LIABILITY', 'SALVAGE VALUE', 'Based on the information', 'Page 3 of 8', 'The damage to insured', 'EVIDENCE', 'The alleged', 'ESTIMATED LOSS', 'ASSOCIATED SURVEYORS'], 'Cause of Loss': ['09)','08)','9)','Liability', 'Loss/Survey Location', '(Attached', '(6) Adequacy'],
                                 'SUPPOSED CAUSE, AS': {'REPORTED'}, 'Cause of loss/ Perils': ['Name of Applicant'], 'SITUATION OF LOSS': ['Regd. Office'],
                                 'The cause of loss is reported': ['In this regard the insured'], 'Cause Of Loss': ['Loss Address']},
               'Cause of Loss (Sentence) Description': {'Our Observation':['09)'],'Failure –':['Incident Details'],'89376':['This report is being'],'CIRCUMSTANCES OF INCIDENT': ['CAUSE OF LOSS', 'Regd. Office :910-911', 'Mumbai Office', 'APPARENT CAUSE'],
                                                        'CIRCUMSTANCES OF LOSS': ['5.00 FIRE BRIGDE REPORT','Statement of Insured'],
                                                        'CIRCUMSTANCES OF': ['CAUSE OF LOSS', 'Mumbai Office'], 'CIRCUMSTA}{CES': ['CAUSE OF LOSS'],
                                                        'Circumstances of': ['Nature & Extent'], 'OCCURRENCE:': ['Page 3 of']},
               'Estimate of Loss': {'88,20,549':['68,11,316'],'final claim is worked out at about': [', as per'],'Insured revised their claim at': ['as under'], 'Estimated Loss': ['Rough Loss Assessment', 'MATERIAL DAMAGE'], 'INSURED’S FINAL CLAIM': ['Insured has not'],
                                    'Insured submitted their final claim bill': ['for their loss'],'INSURED FINAL CLAIM': ['Insured submitted'],
                                    'INSURED FINAL': ['CLAIM'],
                                    'Total Amount Claimed': {'LOSS ASSESSMENT'}, 'Insured’s Claim': {'LOSS ASSESSMENT'},
                                    'PRESENTLY ESTIMATED': {'LOSS RESERVE'}, 'Estimate of Loss': {'ABOUT THE OCCURRENCE'},
                                    "Insured's estimate" : ['ABOUT MAHARASHTRA FLOODS'], 'INSURED’S INTIMATED': ['GROSS LOSS ASSESSED'],
                                    'Claim Amount': ['WebURN Number'], 'Loss Reserve': ['Liability'],
                                    'INSURED’S CLAIM': ['Copy of reinstatement', 'Copy ofquotation', '(Annexure-2)'], 'estimate of loss': ['Amount claimed', 'However, as'],
                                    'Insured has provided us claim bill of': ['for reinstatement'], 'FINAL CLAIM AMOUNT': ['Insured has not'],
                                    'Insured informed that according to them the loss value would be around': ['POLICE REPORT'], 'The insured has claimed': ['Our assessment']},
               'Loss Location (City)': {' Road,':[', Maharashtra'],'Baisara,':[', District'],'Jamnagar Road,':['Insured','. The relevant'],'Distt.':[', Odisha'],'City':['500']},
               'Loss Location (Full) Address': {'Insured :':['Insurer'],'26.12.2017 at':['. The relevant'],'Address':['Location'],
                   'The loss was occurred at the store of the insured located at': ['Affected Property'], 'LOSS LOCATION': ['DATE OF LOSS', 'INSURANCE COVERAGE', 'CAUSE OF LOSS', '(Supervision cameras', 'DATE & TIME', 'SUM INSURED', 'OCCUPATION'], 'LOCATION OF LOSS': ['PERIOD OF INSURANCE', 'INSURANCE COVER',
                                                                                                        'OCCUPATION (AS PER POLICY)', 'OCCUPATION','DATE & TIME',
                                                                                                                              'ADDRESS OF RISK'],
                                                'Loss/Survey Location': {'Damaged Items'}, 'Location of survey': ['Items under'], 'PLACE OF SURVEY': ['WHETHER INCIDENCE', '3.04 CONTACT PERSON'],
                                                'Risk Location':['Period','Sum Insured', 'Occupancy', '5olo of'], 'AFFECTED LOCATION': {'4.00 SURVEY PARTICULARS'},
                                                'Location of loss': 'Loss Description', 'ADDRESS OF LOSS': ['DATE', 'OCCUPATION', 'DESCRIPTION'],
                                                'Site Name': ['Incident Date &'], 'RISK LOCATION OF THE': ['LOSS LOCATION'],
                                                'reported to have occurred (location': ['insured under'], 'Loss Address1': ['Loss Country'], 'Location of the': ['Risk']},
               'Date of Survey': {'us on':['at insured'],'surveyed the loss of / damage on':['to building'],'visited both the loss locations on': {', as per their'}, 'visited the loss location on': {', as per their'}, 'DATE OF SURVEY / VISIT': ['SITUATION OF LOSS', 'REASON FOR DELAY', 'PERSON CONTACTED'],
                                  'Date of Survey': ['Site location','Premises','Cause of Loss', 'and subsequently', 'Place of Survey', 'Special', '& subsequently', 'Property Damaged'],
                                  'DATE OF SURVEY': ['SITUATION OF LOSS', 'TYPE OF POLICY', 'CIRCUMSTANCES OF LOSS', 'CAUSE OF LOSS', 'TYPE OF LOSS','LOSS LOCATION', 'PERSON CONTACTED','CIRCUMSTA}{CES', 'INSTRUCTION', 'CONTACT PERSON', 'REASON FOR DELAY', '4.03', '& subsequently'],
                                   'TIME & DATE OF VISIT': {'REASON FOR DELAY', 'CONTACT PERSON'}, 'Date of survey': ['Cause of loss/ Perils'],
                                  'Date of Visit': ['Reas','Survey','Source of', 'Peril/Cause', 'and subsequent day', 'Regd. Office'], 'DATE OF VISIT /': ['DESCR', "20'h May", 'SURVEY'], 'DATE OF VISIT': ['DESCR', "20'h May", 'CONTACT PERSON'],'DATE & TIME OF SURVEY': ['PLACE OF SURVEY'],
                                  'surveyed the captioned loss': ['in the presence'], 'DATE OF APPLICATION': ['at around'], 'our survey dated': ['on the', 'in connection'], 'the survey was conducted on': ['At the time']
                                  ,'Survey':['Loss']},
               'Final Amount': {'79,58,454':['The above'],'Estimate of Loss : Rs.':['Subject Matter'],'an amount of Rs. Rs.':['/-'],'Total =':['67892-Interim'],'Net Adjusted Loss (r/o)':['14.00 BUSINESS'],'Net Adjusted Loss (A+B)': {'RECOMMENDATION'}, 'Net Adjusted Loss': ['Regd. Office', 'CONCLUSION','14. CONCLUDING REMARKS',
                                                                                                     'Obsolescence', 'Concurrence',
                                                                                                     'CONCURRENCE', 'For Details', 'Note: The',
                                                                                                     'CURRENT ACTION', 'C0NCIIRRANCE','Say..', '11. CONCLUDING REMARKS', 'Less: Excess', '01. INSURANCE', '13. CONCLUDING REMARKS'],
                                'Net loss Assessed': ['Rounded OFF To'], 'Net Loss Value': ['Reinstatement Premium'], 'Net amount recommended for on account payment': ['BASIS OF CALCULATION'],
                                'NET LOSS ASSESSED': ['For Further details'], 'Net Loss Payable': ['Insured consent letter'], 'Net Loss assessed': ['Add: Removal'],
                                'net assessed loss': ['DEATIL OF ASSESSED', '(Rs.'], 'NET ADJUSTED CLAIM': ['For further details', 'CONCURRENCE', 'CONTRIBUTION', 'For Further'], 'NET ADJUSTED': ['CLAIM'],
                                'Net Payable Amount': ['Page 6 of'], 'Nett Loss Assessed': ['Registered office'], 'Net Assessed Loss': ['14.02.13 ADEQUACY']},
               'Gross Loss': {'Assessed = Rs.':['/-'],'Total Sum Insured Rs.':['Visit Details'],'an amount of Rs. Rs.':['/-'],'Total =':['67892-Interim'],'Net Adjusted Loss (r/o)':['14.00 BUSINESS'],'GROSS LOSS ASSESSED': ['Based on inspection','(Based on information', 'NET ADJUSTED CLAI', 'For Further details', 'For further details','NET ADJUSTED LOSS', 'Based on Inspection', 'Based on verification'],
                              'Gross Loss Assessed (A+B)': {'Technological'}, 'Gross Loss Assessed (As per Annexure-A)': ['Less: Retention by'], 'Gross Loss Assessed': ['Technological', 'Less: Depreciation', 'Less Price Variation', 'Net Adjusted Loss'], 'Gross Assessed Loss': ['Less: Depreciation'], 'Total Loss Assessed': ['Less Salvage'],
                              'loss assessed': ['Policy Excess'], 'Gross Loss Claimed and Assessed': ['Less: Min. Policy'],
                              'Gross Loss @ 5.25% (For details, please refer to Annexure - VI)': ['100906.97'], 'Gross Loss': ['Less: Obsolescence'],
                              'Gross Claim Amt': ['Less: Standard Deduc'], 'Gross Assessed loss': ['Less TCS 1%'], 'calculated the gross loss to': ['SALVAGE'],
                              'Gross assessed Loss': ['Less: Depreciation'], 'Gross Total =': ['Less: Depreciation']},
               'Salvage': {'Less Salvage Value =':['Assessed'],'Salvage Value    :':['Reserve'],' Excess clause, we recommend that an amount of Rs.':['/- ('],'had reinstated the property.':['We obtained the'],'Less: Salvage (Refer Para 10)': ['Net off Salvage'], 'Less: Salvage (Refer Para 7)': ['Net Off Salvage'], 'Less: Salvage Value (Net of GST)': ['10,08,59,917.00'],
                           'Less: Salvage Value': ['Subtotal'], 'Less: Salvage =': ['Less: Under Insurance'],
                           'Less: Salvage': ['Net Loss Assessed', 'Net Assessed Loss'], 'Less Salvage (Notional) (Para: 10)': ['Net off Salvage'],
                           'Less Salvage': ['Less Excess', '7 Net Assessed Loss', 'Net Loss Assessed'],
                           'SALVAGE CONSIDERED':['For Further details'], 'agreed to retain the salvage': ['which is around'],'SALVAGE': ['The damaged Stock', 'The Damaged Stock'],
                           'Less for Salvage': ['Value After Salvage'], 'Less:Salvage': ['Sub Total'], 'Less salvage': ['749504']},
               'Under Insurance': {'Under Insurance is':['Based on'],'(Karnataka) - 560034':['Preamble'],'Total= 1,49,00,27,41,000':['12)   Provisional'],'insurance to the extent of':['13.00 ASSESSMENT'],'Less:Under Insurance@8.94%': ['Sub Total'], 'Less: Underinsurance': {'Adjusted Loss'}, 'Less Under Insurance (Para: 11.2)': ['Net off Under Insurance'],
                                   'Less Under Insurance': ['Adjusted Loss'], 'Less: Under Insurance (Refer Para 11.2)': ['Net off Under'], 'Less: Under Insurance (Refer Para 8.2)': ['Net Off Under'],'Less: Under Insurance': ['Amount', 'Adjusted Loss', 'Subtotal', 'Adj usted Loss'], 'Under Insurance': ['Thus', '14.2 P&M'],
                                   'Less: UI': ['Adjusted Loss'], 'UNDER INSURANCE': ['Insured had provided us', 'For Further details', 'For further details', 'For Stock', 'We', 'Insured has provided'],
                                   'less for Under Insurance': ['Value After Under Insurance'], 'Underinsurancc': ['Adjusted Loss']},
               'Excess': {'Policy Excess =':['/-'],'Less Policy Excess @ 5% of Claim Amount subject to minimum of Rs. 5 Lac =':['Net Liability'],'a minimum of INR':['/- each'],' Excess clause, we recommend that an amount of Rs.':['/- ('],'Less: Excess (Refer Para 11.3)': ['Net Adjusted Loss'], 'Less: Excess 5% of the claim amount subject to a': ['minimum of'], 'Less: Excess': ['Net Adjusted Loss', 'standard', 'LOSS ADJUSTED'], 'Less:Policy Excess': ['Assessed Loss'], 'Policy Excess': ['Net Loss Value', 'Net Loss Payable'],
                          'Less Excess': ['Net loss Assessed', 'Net Adjusted Loss'], 'Less Excess (Para: 11.3)': ['Net Adjusted Loss'],
                          'EXCESS': ['Excess for location Maharashtra', '5% of claim amount', '5% of Claim Amount'], 'NET ASSESSED LOSS': ['UNDER INSURANCE'],
                          'Less for Policy Excess': ['Loss After Policy Excess']},
               'Depreciation': {'actual invoices.':['There'],'Less: Depreciation (Refer Para 9)': ['Net off Depreciation'], 'Less: Depreciation (Refer Para 6)': ['Net Off Depreciation'], 'Less Depreciation @ 20%': ['756034'],
                                'Less: Depreciation': ['Loss Assessed', 'Less: Salvage', 'Net Adjusted Loss'], 'deprecition': ['Net Loss assessed'],
                                'Less Depreciation @4% per year for 1 year': ['Sub Total'], 'Less Depreciation (Para: 9)': ['Net off Depreciation'], 'Less Depreciation': ['Less Salvage'], 'DEPRICIATION': ['For Further details', 'Policy is'], 'DEPRECIATION': ['Based on','Policy is on', 'OBSOLESCENCE', 'MARKET VALUE']},
               'Assessment (Full)': {'following observations were made;': {'Regd. Office :'}, 'OUR SURVEY/ OBSERVATION:': ['SAPIENT INSURANCE']},
               'Property Damaged': {},

               'Total Amount_ASS': {}, #don't require

               'Total Amount': {},  #don't require
               'Debris Removal':{'As per policy, the insured have taken coverage of Rs.':['/-']}

               }
    out = {'IL Claim Ref No': None, 'Date of Loss': None, 'Cause of Loss': None,
           'Cause of Loss (Sentence) Description': None, 'Estimate of Loss': None, 'Loss Location (City)': None,
           'Loss Location (Full) Address': None, 'Date of Survey': None, 'Property Damaged_ASS': None,
           'Quantity_ASS': None, 'Unit_ASS': None, 'Amount per Quantity_ASS': None, 'Total Amount_ASS': None,
           'Property Damaged': None, 'Quantity': None, 'Unit': None, 'Amount per Quantity': None, 'Total Amount': None,
           'Gross Loss': None, 'Depreciation': None, 'Excess': None, 'Salvage': None, 'Debris Removal': None,
           'Under Insurance': None, 'Final Amount': None, 'Assessment (Full)': None, 'Remarks': None}
    # out={}
    exu = 0


    def ext1(str1):
        tmp = ''


        tt = []
        tB = str1
        tB = re.sub(' +', ' ', tB)
        tB = tB.replace(',', '')
        tC = []
        for a in tB.split():
            a = a.title().replace('Rs.', '').replace('Rs', '').strip()
            if a.count('.') > 1 or len(a.title().replace('Rs', '').replace('.', '').replace(',', '').strip()) > 10:
                continue
            a1 = a
            if '.' in a1 and len(a1.split('.')) == 2:
                a1 = re.sub('[\.]', '', a)
            if a1.isdigit() == True:
                if len(tC) == 0:
                    tC.append(float(a))
                else:
                    if float(a) > tC[0] or float(a) == tC[0]:
                        tC = []
                        tC.append(float(a))
        if len(tC) > 0:
            tmp = tC[0]
            tmp = str(tmp)
        if len(tmp) > 2 and '.' in tmp:
            if int(tmp.split('.')[-1]) == 0:
                tmp = tmp.split('.')[0]
        return tmp





    for k1 in dic_key:
        exu = 0
        for k2 in dic_key[k1]:
            vs = []
            for x in dic_key[k1][k2]:

                # print(k1, k2, x)
                if x and k2 in txt:


                    clm_no = ((txt.split(k2)[1]).split(x)[0]).strip()

                    if k1 == 'Date of Loss':
                        # print(k1, k2)
                        if ':' in clm_no:
                            clm_no = clm_no.split(':')[1].strip()
                            clm_no = clm_no.replace(':', '').replace(') Insurers', '').strip()
                        if 'THE INCIDENT' in clm_no:
                            clm_no = clm_no.replace('THE INCIDENT', '').strip()
                        if '1.7' in clm_no:
                            clm_no = clm_no.split('1.7')[0].strip()
                        else:
                            clm_no = clm_no.replace('@ night hours of Tuesday dated','').replace('Reportedly on','').\
                                replace('Name of Not applicable.','').replace('As reported by the Insured on','').replace('3.2','').\
                                replace('Reportedly', '').replace(',', '').strip()

                        if k1 == 'IL Claim Ref No':
                            temp = re.findall("\w{2,3}\d{7,10}", clm_no)
                            if len(temp) >= 1:
                                clm_no = temp[0].replace('\'', '').replace(':','').strip().split(' ')[0]
                                temp = clm_no
                                if temp.isdigit():
                                    clm_no = ''
                            else:
                                clm_no = ''

                        # else:
                        #     # clm_no = clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        #     clm_no = clm_no.replace(':', '').replace('-','').strip()
                        #     # print(clm_no)
                            # print(clm_no)

                    if k1 == 'Cause of Loss':
                        # print(clm_no)
                        clm_no = clm_no.replace('5.1', '').replace(':','').replace('4.1','').\
                            replace('-', '').replace('6.2.1', '').replace('9.01', '').replace('10.01', '').strip()
                        # if 'AND' in clm_no:
                        #     clm_no = clm_no.split('AND')[1:]
                        # clm_no = clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        # clm_no = clm_no.replace(':', '').strip()
                        # clm_no = clm_no.rsplit('(', 1)[0].strip()

                    if k1 == 'Cause of Loss (Sentence) Description':
                        clm_no = clm_no.replace('\uf0b7','').strip()
                        # print(clm_no)
                        if 'IAR' and 'of' in clm_no:
                            temp = clm_no[clm_no.find(' IAR'):clm_no.find('of ') + 5]
                            clm_no = clm_no.replace(temp, '').strip()
                        clm_no = clm_no.rsplit('(', 1)[0].strip().replace(':', '')

                    if 'OUR SURVEY' in clm_no:
                        clm_no = clm_no.split('OUR SURVEY')[0]
                    if 'OUR SURVEY:' in clm_no:
                        clm_no = clm_no.split('OUR SURVEY:')[0]

                    if k1 == 'Estimate of Loss':
                        if x == 'THE SOLAR PUMP HOUSE':
                            clm_no = clm_no.split('INSURED')[0].strip()

                            # print(clm_no)
                        if 'INR.' in clm_no:
                            clm_no = clm_no.strip().split('INR.')[1].lstrip().split(' ')[0]
                        if 'INR' in clm_no:
                            clm_no = clm_no.strip().split('INR')[1].lstrip().split(' ')[0]
                        if 'Rs.' in clm_no:
                            clm_no = clm_no.strip().split('Rs.')[1].lstrip().split(' ')[0]
                            # print(clm_no)
                        if '₹' in clm_no:
                            clm_no = clm_no.strip().split('₹')[1].lstrip().split(' ')[0].replace('/', '').replace('-',
                                                                                                                  '').replace(
                                '.', '')
                            # print(clm_no)

                        else:

                            # clm_no = clm_no.split(clm_no.split(' ')[-1])[0].strip()
                            clm_no = clm_no.replace('/','').replace('-','').replace('Nature Of Loss 100310193979105 FIR028962557 HETERO WIND POWER LTD', '').strip()
                            # print(clm_no)

                    if k1 == 'Loss Location (Full) Address':
                        clm_no = clm_no.replace('Address', '').replace(':', '').\
                            replace('TIME &', '').replace('LOCATION, WITH PIN CODE', '').\
                            replace('As per','').replace('Policy','').replace('Loss 2', '').replace('Loss 3','').\
                            replace('Loss City','').replace('Loss State', '').replace('Loss Pincode', '').replace('-', '').\
                            replace('  ', ' ').replace('3.5', '').strip()
                        # if 'Policy' in clm_no:
                        #     clm_no = (clm_no.split('Policy')[0]).strip()
                        #     clm_no = clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        # if 'Section wise' in clm_no:
                        #     clm_no = (clm_no.split('Section wise')[0]).strip()
                        # if 'insured\'s' in clm_no and 'premises' in clm_no:
                        #     clm_no = ''
                        #     clm_no = ((txt.split('POLICY PARTICULARS')[1]).split('POLICY NO.')[0]).strip()
                        #     clm_no = clm_no.split('INSURED')[1].strip()
                        # if 'IAR' and 'of' in clm_no:
                        #     temp = clm_no[clm_no.find(' IAR'):clm_no.find('of ') + 5]
                        #     clm_no = clm_no.replace(temp, '').strip()

                    if k1 == 'Date of Survey':
                        # print(clm_no)
                        if x == 'PLACE OF SURVEY':
                            clm_no = clm_no.strip()

                        else:
                            clm_no = clm_no.replace('on','').replace(',','').replace('/ VISIT', '').replace(':','').replace('&','').strip()
                            # print(clm_no)
                            # if len(clm_no.split(' ')) > 1 and ':' in clm_no:
                            #     clm_no = clm_no.strip().split(' ')[1].strip()
                            #     if 'Loss' in clm_no:
                            #         clm_no = (clm_no.split('Loss')[0]).strip()
                            #         clm_no = clm_no.split(clm_no.split(' ')[-1])[0].strip()
                            #         # print(clm_no)
                            # if len(clm_no.split(' ')) > 1 and ':' not in clm_no:
                            #     clm_no = clm_no.strip()#.split(' ')
                            #     print(clm_no)
                                # list1 = list1[0]
                                # print(clm_no)
                                # print(list1[0])
                                # clm_no = clm_no[0] + clm_no[1] + clm_no[2]

                    if k1 == 'Final Amount':
                        #                         clm_no=clm_no.split(clm_no.split(' ')[-1])[0].strip()
                        # clm_no = (
                        # (clm_no.split(' ' + clm_no.strip().split(' ')[-1] + ' ')[0].strip()).rsplit(' ', 1)[0]).strip()
                        clm_no = clm_no.replace('*', '').replace('=', '').replace('for', '').strip()
                        # print(clm_no)
                        # if 'INR' in clm_no:
                        #     clm_no = clm_no.strip().split('INR')[1].lstrip().split(' ')[0]
                        # if 'Rs.' in clm_no:
                        #     clm_no = clm_no.strip().split('Rs.')[1].lstrip().split(' ')[0]
                        # if '₹' in clm_no:
                        #     clm_no = ''
                        #     clm_no = ((txt.split('Net Assessed Loss')[1]).split('Annexure')[0]).strip()
                        #     clm_no = clm_no.strip().split('₹')[1].lstrip().split(' ')[0].replace('/', '').replace('-',
                        #                                                                                           '').replace(
                        #         '.', '')
                        temp = re.findall("\d+,?\d+.?\d+?", clm_no)
                        # if len(temp) > 0:
                        #     clm_no = temp[0].split(' ')[0]
                        # print(temp)

                    if k1 == 'Gross Loss':

                        # clm_no = (
                        # (clm_no.split(' ' + clm_no.strip().split(' ')[-1] + ' ')[0].strip()).rsplit(' ', 1)[0]).strip()
                        # print(clm_no)
                        if len(clm_no.split(' ')) > 1:
                            clm_no = clm_no.split()[0].strip()
                            # print(clm_no)
                        # print(clm_no)
                        if 'INR' in clm_no:
                            clm_no = clm_no.split('INR')[1].lstrip().split(' ')[0]
                        elif 'Rs.' in clm_no:
                            clm_no = clm_no.split('Rs.')[1].lstrip().split(' ')[0]
                            # print(clm_no)
                        else:
                            clm_no = clm_no.split(' ')[0].strip()
                            # print(clm_no)

                    if k1 == 'Salvage':
                        clm_no = clm_no.strip()
                        # print(clm_no)
                        # if '%' in clm_no:
                        #     clm_no = clm_no.split()[1].strip()
                        #     print(clm_no)
                            # print(clm_no)
                        # print(clm_no)
                        if 'Rs.' in clm_no:
                            clm_no = clm_no.split('Rs.')[1].lstrip().split(' ')[0]
                            # print(clm_no)
                        elif 'INR' in clm_no:
                            clm_no = clm_no.split('INR')[1].lstrip().split(' ')[0]
                        if 'Value' in clm_no:
                            clm_no = clm_no.split('Value')[1].lstrip().split(' ')[0]
                        if len(clm_no.split(' ')) > 1:
                            clm_no = clm_no.split()[-1].strip()
                            # print(len(clm_no.split(' ')))

                    #                         if clm_no!='NA':
                    #                             if 'Rs.' in clm_no:
                    #                                 temp=re.findall("Rs..*\d+,?.?\d+.?\d*-?", clm_no)
                    #                                 if len(temp)>=1:
                    #                                     clm_no=temp[0]
                    #                             elif 'INR' in clm_no:
                    #                                 temp=re.findall("INR.*\d+,?.?\d+.?\d*", clm_no)
                    #                                 if len(temp)>=1:
                    #                                     clm_no=temp[0]

                    if k1 == 'Under Insurance':
                        clm_no = clm_no.replace('-','').replace('=', '').strip()
                        # print(clm_no)
                        # print(clm_no)
                        # if len(clm_no.split(' ')) > 1:
                        #     clm_no = clm_no.split()[-1].strip()
                            # print(clm_no)
                        # clm_no = clm_no.split(' ')[-1].strip()
                        # print(clm_no)
                        # if '%' in clm_no:
                        #     clm_no = clm_no.split()[1].strip()
                            # print(clm_no)

                    if k1 == 'Excess':
                        clm_no = clm_no.replace('@ 5%','').strip()
                        if len(clm_no.split()) > 1:
                            clm_no = clm_no[-1]
                        # print(clm_no)
                        if 'Rs.' in clm_no:
                            clm_no = clm_no.split('Rs.')[1].strip()
                            temp = re.findall("\d+,?.?.*?\d+.?\d*-?", clm_no)
                            if len(temp) >= 1:
                                clm_no = temp[0]
                        elif 'INR' in clm_no:
                            clm_no = clm_no.split('INR')[1].strip()
                            temp = re.findall("\d+,?.?.*?\d+.?\d*", clm_no)
                            if len(temp) >= 1:
                                clm_no = temp[0]
                        if '₹' in clm_no:
                            clm_no = clm_no.split('₹')[1].strip()
                            temp = re.findall("\d+,?.?.*?\d+.?\d*", clm_no)
                            if len(temp) >= 1:
                                clm_no = temp[0]
                                # print(clm_no)
                    #                         if 'Rs.' in clm_no:
                    #                             clm_no=clm_no.split('Rs.')[1].lstrip().split(' ')[0]
                    #                         if 'INR' in clm_no:
                    #                             clm_no=clm_no.split('INR')[1].lstrip().split(' ')[0]
                    #                         if '₹' in clm_no:
                    #                             clm_no=clm_no.strip().split('₹')[1].lstrip().split(' ')[0].replace('/','').replace('-','').replace('.','')

                    if k1 == 'Depreciation':
                        # print(clm_no)
                        if '%' in clm_no:
                            clm_no = (clm_no.split('%')[1].strip()).split()[-1]
                            # print(clm_no)

                        else:
                            # clm_no = clm_no.split()[-1].strip()
                            if len(clm_no.split(' ')) > 1:
                                clm_no = clm_no.split()[-1].strip()

                            # print(clm_no)
                        # if 'Depreciation' in clm_no:
                        #     clm_no = clm_no.strip().split('Depreciation')[1].strip().split(' ')[0]
                        #     # print(clm_no)
                        # else:
                        #     temp = re.findall("\d+,\d+.?\d*", clm_no)
                        #     if len(temp) >= 1:
                        #         clm_no = temp[0].split(' ')[0]
                        #         print(clm_no)

                    if k1 == 'Total Amount_ASS':
                        clm_no = clm_no.strip()
                    if k1 == 'Total Amount':
                        clm_no = clm_no.strip()

                    if k1 == 'Assessment (Full)':
                        if 'IAR' and 'of' in clm_no:
                            temp = clm_no[clm_no.find(' IAR'):clm_no.find('of ') + 5]
                            clm_no = clm_no.replace(temp, '').strip()
                        if '6.00' in clm_no:
                            clm_no = clm_no.replace('6.00', '').strip()
                        else:
                            clm_no = clm_no.strip()
                            clm_no = clm_no.replace('\uf0b7', '').strip()
                            # print(clm_no)

                    if k1 == 'Property Damaged_ASS':
                        clm_no = clm_no.strip()

                    if k1 == 'Property Damaged':
                        clm_no = clm_no.replace('IAR-1805-17396_Reliance Retail Ltd_FSR', '').replace('Page 3 of 4',
                                                                                                      '').strip()
                    #                         clm_no=clm_no.split(clm_no.split(' ')[-1])[0].strip()
                    #                         if 'Date' in clm_no:
                    #                             clm_no=''
                    #                             clm_no=txt.split('Cause of Loss')[0]
                    #                             if 'Cash Loss' not in clm_no:
                    #                                 clm_no=(clm_no.rsplit('Loss',1)[1]).strip()
                    #                                 clm_no=clm_no.split(clm_no.split(' ')[-1])[0].strip()
                    #                             elif 'Cash Loss' in clm_no:
                    #                                 clm_no='Cash Loss'
                    #                         c_new=[]
                    #                         c_new_item=[]
                    #                         c_new_amt=[]
                    #                         for r in re.split('\d+\s',clm_no):
                    #                             if len(r)==0:
                    #                                 continue
                    #                             c_new.append(r)
                    #                         for r in range(len(c_new)):
                    #                             if r%2==0:
                    #                                 c_new_item.append(c_new[r])
                    #                             else:
                    #                                 c_new_amt.append(c_new[r])
                    #                         item=''
                    #                         amt=''
                    #                         for c in c_new_item:
                    #                             item +=c +', '
                    #                         item=item.strip(', ')
                    #                         for c in c_new_amt:
                    #                             amt +=c +', '
                    #                         amt=amt.strip(', ')
                    #                         clm_no=item
                    if k1 == 'Amount per Quantity':
                        c_new = []
                        c_new_item = []
                        c_new_amt = []
                        for r in re.split('\d+\s', clm_no):
                            if len(r) == 0:
                                continue
                            c_new.append(r)
                        for r in range(len(c_new)):
                            if r % 2 == 0:
                                c_new_item.append(c_new[r])
                            else:
                                c_new_amt.append(c_new[r])
                        item = ''
                        amt = ''
                        for c in c_new_item:
                            item += c + ', '
                        item = item.strip(', ')
                        for c in c_new_amt:
                            amt += c + ', '
                        amt = amt.strip(', ')
                        clm_no = amt
                    if len(vs) == 0:
                        vs.append(len(clm_no))
                        vs.append(clm_no)
                        out[k1] = vs[1]
                        exu = 1
                        continue
                    if len(vs) != 0:
                        if len(clm_no) > vs[0]:
                            continue
                        vs = []
                        vs.append(len(clm_no))
                        vs.append(clm_no)
                        out[k1] = vs[1]
                        exu = 1

            if exu == 1:
                break
            #         out[k1] = clm_no
            #
            #         exu = 1
            #         break
            # if exu == 1:
            #     break
    print(fl, '#########################################################################')

    doc_num = ''
    new_link = ''
    if out['IL Claim Ref No'] is not None:
        doc_num = out['IL Claim Ref No']
    if len(doc_num) > 5:
        doc_num = re.sub('[^A-Z0-9]', '', doc_num)
        # doc_num=
        print(doc_num, 'lllll')
    for d in data:
        if doc_num == d.split('&$')[0]:
            print('yyyyyyyyyy', d.split('&$')[2])
            new_link = d.split('&$')[2]
            # new_link='new_link'
            break
    for d in data_one:
        if doc_num == d.split('&$')[0]:
            # out['Estimate of Loss']="No Claim"
            out['Gross Loss'] = "Nil"
            out['Salvage'] = "Nil"
            out['Excess'] = "Nil"
            out['Final Amount'] = "Nil"
            out['Under Insurance'] = "Nil"
            out['Depreciation'] = "Nil"
            out['Debris Removal'] = "Nil"

    print(out)
    mbk = openpyxl.load_workbook(path, read_only=False)
    sht = mbk.get_sheet_by_name('Sheet1')
    p = 3 + cnt
    cnt += 1
    sht.cell(row=p, column=1).value = "Pro"
    # sht.cell(row=p, column=2).value = fl
    sht.cell(row=p, column=2).hyperlink=fl
    sht.cell(row=p, column=2).value=fl.split('.txt')[0]
    c = 3
    for t in out:
        sht.cell(row=p, column=c).value = out[t]
        c += 1
    mbk.save(path)
