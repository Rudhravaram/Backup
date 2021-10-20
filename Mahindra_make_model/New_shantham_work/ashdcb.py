import re
import os
import subprocess
from openpyxl import load_workbook
from shutil import copyfile,copy2
import xlrd
import pandas as pd
import glob
import shutil
from cryptography.fernet import Fernet
from xlutils.copy import copy
# from Icici_lombard_mail.Packages.xlwings_doc import xlwings as xw
import xlwings as xw
import datetime

file='Sample_ - Copy - Copy (10) - Copy.xlsx'

wb_new_myra= load_workbook(filename=file)
wb_orig_new_myra = xlrd.open_workbook(file)
print(wb_new_myra.sheetnames)

for num_sheet in wb_new_myra.sheetnames:
    if num_sheet == 'GHSJ':
        ws_new_myra = wb_new_myra[num_sheet]
        sheet_new_myra = wb_orig_new_myra.sheet_by_name(num_sheet)
    elif num_sheet == 'To be Mapp - set 2':
        ws_set1_sheet = wb_new_myra[num_sheet]
        sheet_set1_sheet = wb_orig_new_myra.sheet_by_name(num_sheet)
if ws_new_myra:
    List_=[1]
    va=''
    for i in List_:
        tem_t=False
        if not va:
            tem_t=True
            va = str(i)
        if va and not tem_t:
            va=va+','+str(i)
    print(va)
    ws_new_myra.cell(row=2, column=2).value=va
wb_new_myra.save(filename=file)
wb_new_myra.close()

