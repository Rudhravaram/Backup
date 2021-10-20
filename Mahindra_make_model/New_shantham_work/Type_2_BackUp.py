import re
import os
import subprocess
from openpyxl import load_workbook
from shutil import copyfile,copy2
import xlrd
import pandas as pd
import glob
import shutil
from cryptography.fernet import Fernet
from xlutils.copy import copy
# from Icici_lombard_mail.Packages.xlwings_doc import xlwings as xw
import xlwings as xw
import datetime

file='Sample.xlsx'

wb_new_myra= load_workbook(filename=file)
wb_orig_new_myra = xlrd.open_workbook(file)
print(wb_new_myra.sheetnames)

for num_sheet in wb_new_myra.sheetnames:
    if num_sheet == 'Base':
        ws_new_myra = wb_new_myra[num_sheet]
        sheet_new_myra = wb_orig_new_myra.sheet_by_name(num_sheet)
    elif num_sheet == 'To be Mapp - set 1':
        ws_set1_sheet = wb_new_myra[num_sheet]
        sheet_set1_sheet = wb_orig_new_myra.sheet_by_name(num_sheet)


if ws_new_myra :
    start_column_ws = 2
    start_column_ws2 = 3
    start_Varient = 4
    Fuel_Type= 8
    fuel_type2=10
    Set1_id = 9
    Set2_id = 11
    Per = 10  # 12
    Id = 1
    start_row_ws_i = 2
    for i in range(sheet_new_myra.nrows):
        temp_Make_Name = ws_new_myra.cell(row=start_row_ws_i, column=start_column_ws).value
        temp_Model_Name = ws_new_myra.cell(row=start_row_ws_i, column=start_column_ws2).value
        temp_Variant_Name = ws_new_myra.cell(row=start_row_ws_i, column=start_Varient).value
        temp_Fuel_Type = ws_new_myra.cell(row=start_row_ws_i, column=Fuel_Type).value
        temp_Set1_id = ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value
        start_row_ws=2
        List_fi=[]
        List_fi2=[]
        List_fi1=[]
        List_fi_prob=[]
        List_fi21=[]
        DirectMatc=[]
        LessProb=[]
        for j in range(sheet_set1_sheet.nrows):
            print(start_row_ws_i,start_row_ws)
            temp_MAKE_NAME=ws_set1_sheet.cell(row=start_row_ws, column=start_column_ws).value
            temp_MODEL_NAME = ws_set1_sheet.cell(row=start_row_ws, column=start_column_ws2).value
            temp_VARIANT_NAME = ws_set1_sheet.cell(row=start_row_ws, column=start_Varient).value
            temp_Fuel_Type2= ws_set1_sheet.cell(row=start_row_ws, column=Fuel_Type).value
            temp_Id= ws_set1_sheet.cell(row=start_row_ws, column=Id).value
            print('####################################Start###########################################')
            print(temp_Make_Name)
            print(temp_MAKE_NAME)
            print(temp_Model_Name)
            print(temp_MODEL_NAME)
            print(temp_Variant_Name)
            print(temp_VARIANT_NAME)
            print(temp_Set1_id)
            print(temp_Id)
            print('########################################End##########################################')
            temp_varTrFa=False
            if (str(temp_Make_Name).replace('MAHINDRA SSANGYONG','MAHINDRA').replace('MARUTI SUZUKI','MARUTI').replace('AND','&').replace('MAHINDRA & MAHINDRA','MAHINDRA').lower().strip() == str(temp_MAKE_NAME).replace('AND','&').replace('MAHINDRA & MAHINDRA','MAHINDRA').replace('MARUTI SUZUKI','MARUTI').replace('  ',' ').lower().strip()) and (str(temp_Model_Name).upper().strip()==str(temp_MODEL_NAME).upper().strip()):
                if (str(temp_VARIANT_NAME).replace('.','').replace('-','').upper().replace('BHARAT', 'BS').replace('   ', ' ').strip() == str(temp_Variant_Name).replace('.','').replace('-','').upper().strip()) and (str(temp_Fuel_Type2).replace('.','').replace('-','').lower().strip().__contains__(str(temp_Fuel_Type).replace('.','').replace('-','').lower().strip())):
                    print('jshbndhdhhhehhed')
                    print('sdg')
                    print('@@@@@@@@@@@@@@@@@@@@@@@Start@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
                    print(Set1_id)
                    print(Id)
                    print('@@@@@@@@@@@@@@@@@@@@@@@@End@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
                    te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                    if te_id not in DirectMatc:
                        DirectMatc.append(te_id)
                    List_fi=[]
                    List_fi2 = []
                    List_fiContains=[]
                    break
                elif (str(temp_Variant_Name).replace('.','').replace('-','').upper().replace('BHARAT', 'BS').replace('   ', ' ').strip().__contains__(str(temp_VARIANT_NAME).replace('.','').replace('-','').upper().strip())) and (str(temp_Fuel_Type2).replace('.','').replace('-','').lower().strip().__contains__(str(temp_Fuel_Type).replace('.','').replace('-','').lower().strip())):
                    print('s,sdhcjhagcjgsjhcdshdchjsjdc')
                    te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                    if te_id not in List_fi_prob:
                        List_fi_prob.append(te_id)
                elif (str(temp_VARIANT_NAME).replace('.','').replace('-','').upper().replace('BHARAT', 'BS').replace('   ', ' ').strip().__contains__(str(temp_Variant_Name).replace('.','').replace('-','').upper().strip())) and (str(temp_Fuel_Type2).replace('.','').replace('-','').lower().strip().__contains__(str(temp_Fuel_Type).replace('.','').replace('-','').lower().strip())):
                    print('ssdkhbnadjna')
                    te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                    if te_id not in List_fi_prob:
                        List_fi_prob.append(te_id)
                else:
                    print('ssdkhassdbnadjna')
                    temp_Var = str(temp_Variant_Name).replace('.','').replace('-','').replace('K10','').replace('1.1','').replace('  ','').upper().strip().split(' ')
                    V_J=str(temp_VARIANT_NAME).replace('.','').replace('-','').upper().replace('DI','').replace(' ','').replace('STANDARD', 'STD').strip()
                    temp_ftev=False
                    print('temp_Var',temp_Var)
                    for i in temp_Var:
                        if i.upper().strip().__contains__(V_J):
                            if (str(temp_Fuel_Type2).replace('.','').replace('-','').lower().strip().__contains__(str(temp_Fuel_Type).replace('.','').replace('-','').lower().strip())):
                                temp_ftev=True
                                print('%%%%%%%%%%%%%%%%%%%Start%%%%%%%%%%%%%%%%%%%')
                                print(temp_Set1_id)
                                print(temp_Id)
                                print('%%%%%%%%%%%%%%%%%%%End%%%%%%%%%%%%%%%%%%%')
                                te_id=ws_set1_sheet.cell(row=start_row_ws,column=Id).value
                                if te_id not in List_fi:
                                    List_fi.append(te_id)
                    if not temp_ftev:
                        Upperss=False
                        temp_Var = str(temp_Variant_Name).replace('.', '').replace('-', '').replace('K10', '').replace('1.1', '').replace('  ', '').upper().strip()
                        V_J = str(temp_VARIANT_NAME).replace('AT 1.2','').replace('1.2','').replace('1.1', '').replace('.', '').replace('-', '').upper().replace('DI','').replace('STANDARD', 'STD').strip().split(' ')
                        print('sdvkkns')
                        print("V_J",V_J)
                        for i in V_J:
                            print('sdvkknssxf')
                            print(i)
                            if i.upper().strip().__contains__(temp_Var.upper().strip()):
                                Upperss = True
                                print('jhsdgdcjhdcjshjhd')
                                te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                if te_id not in List_fi21:
                                    List_fi21.append(te_id)
                    if not Upperss:
                        print('jsefhgjhgwejfb sdgfhvjsgf')
                        temp_Var = str(temp_Variant_Name).replace('.', '').replace('-', '').replace('K10', '').replace('1.1', '').replace('  ', '').upper().strip().split(' ')
                        V_J = str(temp_VARIANT_NAME).replace('AT 1.2', '').replace('1.2', '').replace('1.1','').replace('.','').replace('-', '').upper().replace('DI', '').replace('STANDARD', 'STD').strip()
                        print("LessProb", V_J)
                        for i in range(len(temp_Var)):
                            print(i)
                            if len(temp_Var) >= 2:
                                if V_J.upper().strip().__contains__(temp_Var[0].upper().strip()) and V_J.upper().strip().__contains__(temp_Var[1].upper().strip()):
                                    te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                    if te_id not in LessProb:
                                        LessProb.append(te_id)
                                else:
                                    if V_J.upper().strip().__contains__(temp_Var[0].upper().strip()):
                                        te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                        if te_id not in LessProb:
                                            LessProb.append(te_id)
                            elif V_J.upper().strip().__contains__(temp_Var[0].upper().strip()):
                                te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                if te_id not in LessProb:
                                    LessProb.append(te_id)
                    if not List_fi and not LessProb:
                        te_ids = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                        if te_ids not in List_fi2:
                            List_fi2.append(te_ids)
            elif (str(temp_Make_Name).replace('MAHINDRA SSANGYONG','MAHINDRA').replace('.','').replace('-','').replace('AND', '&').replace('MARUTI SUZUKI', 'MARUTI').replace('MAHINDRA & MAHINDRA', 'MAHINDRA').lower().strip() == str(temp_MAKE_NAME).replace('.','').replace('-','').replace('AND','&').replace('MAHINDRA & MAHINDRA', 'MAHINDRA').replace('MARUTI SUZUKI', 'MARUTI').replace('  ',' ').lower().strip()) and (str(temp_MODEL_NAME).replace('.','').replace('-','').upper().strip().__contains__(str(temp_Model_Name).replace('.','').replace('-','').upper().strip())):
                print('asdlvjnksjbdvjhbbebfbmadbv,sndvskjdnn,sdn')
                print('sknvknkrvnknerkjvjnvnkjknerjrv')
                temp_getmmod = str(temp_Model_Name).replace('.','').replace('-','').upper().strip()
                temp_getVarf = str(temp_Variant_Name).replace('.','').replace('-','').upper().strip().split(' ')
                for i in temp_getVarf:
                    var_fin = temp_getmmod + ' ' + i
                    if str(var_fin).upper().strip()==str(temp_MODEL_NAME).upper().strip():
                        if (str(temp_Fuel_Type2).replace('.','').replace('-','').lower().strip().__contains__(str(temp_Fuel_Type).replace('.','').replace('-','').lower().strip())):
                            print('sdgkskjgkjs')
                            te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                            if te_id not in List_fi1:
                                List_fi1.append(te_id)
                if not List_fi1:
                    te_ids = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                    if te_ids not in List_fi2:
                        List_fi2.append(te_ids)
                print('Yes')
            start_row_ws = start_row_ws + 1
        if DirectMatc:
            PerCent = '100'
            print("DirectMatc",DirectMatc)
            List_fi=[]
            List_fi2 = []
            List_fi1 = []
            List_fi_prob = []
            va = ''
            for i in DirectMatc:
                tem_t = False
                if not va:
                    tem_t = True
                    va = str(i)
                if va and not tem_t:
                    va = va + ', ' + str(i)
            print(va)
            ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value = va
            ws_new_myra.cell(row=start_row_ws_i, column=Per).value = PerCent
        elif List_fi:
            PerCent = '95'
            print("List_fi")
            List_fi2 = []
            List_fi1=[]
            List_fi_prob=[]
            va=''
            for i in List_fi:
                tem_t = False
                if not va:
                    tem_t = True
                    va = str(i)
                if va and not tem_t:
                    va = va + ', ' + str(i)
            print(va)
            ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value = va
            ws_new_myra.cell(row=start_row_ws_i, column=Per).value = PerCent
        elif List_fi21:
            PerCent = '90'
            print("List_fi")
            List_fi2 = []
            List_fi1 = []
            List_fi_prob = []
            va = ''
            for i in List_fi21:
                tem_t = False
                if not va:
                    tem_t = True
                    va = str(i)
                if va and not tem_t:
                    va = va + ', ' + str(i)
            print(va)
            ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value = va
            ws_new_myra.cell(row=start_row_ws_i, column=Per).value = PerCent
        elif List_fi1:
            PerCent = '90'
            print("List_fi1")
            List_fi2=[]
            print('skejhfkjkjefsw')
            va = ''
            for i in List_fi1:
                tem_ts = False
                if not va:
                    tem_ts = True
                    va = str(i)
                if va and not tem_ts:
                    va = va + ', ' + str(i)
            print(va)
            ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value = va
            ws_new_myra.cell(row=start_row_ws_i, column=Per).value = PerCent
        elif List_fi_prob and not List_fi:
            PerCent = '80'
            print("List_fi_prob")
            List_fi2 = []
            List_fi=[]
            va=''
            for i in List_fi_prob:
                tem_t = False
                if not va:
                    tem_t = True
                    va = str(i)
                if va and not tem_t:
                    va = va + ', ' + str(i)
            print(va)
            ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value = va
            ws_new_myra.cell(row=start_row_ws_i, column=Per).value = PerCent
        elif LessProb:
            PerCent = '75'
            print("LessProb")
            List_fi2 = []
            List_fi1 = []
            List_fi_prob = []
            va = ''
            for i in LessProb:
                tem_t = False
                if not va:
                    tem_t = True
                    va = str(i)
                if va and not tem_t:
                    va = va + ', ' + str(i)
            print(va)
            ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value = va
            ws_new_myra.cell(row=start_row_ws_i, column=Per).value = PerCent
        else:
            if List_fi2:
                PerCent = '50 (Make-Model)'
                print("List_fi2")
                va = ''
                for i in List_fi2:
                    tem_t1 = False
                    if not va:
                        tem_t1 = True
                        va = str(i)
                    if va and not tem_t1:
                        va = va + ', ' + str(i)
                print(va)
                ws_new_myra.cell(row=start_row_ws_i, column=Set1_id).value = va
                ws_new_myra.cell(row=start_row_ws_i, column=Per).value = PerCent
        start_row_ws_i=start_row_ws_i+1
    wb_new_myra.save(filename=file)
    wb_new_myra.close()