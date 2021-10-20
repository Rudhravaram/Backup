import re
import os
import subprocess
from openpyxl import load_workbook
from shutil import copyfile,copy2
import xlrd
import pandas as pd
import glob
import shutil
from cryptography.fernet import Fernet
from xlutils.copy import copy
# from Icici_lombard_mail.Packages.xlwings_doc import xlwings as xw
import xlwings as xw
import datetime
import collections
from fuzzywuzzy import fuzz, process

def get_final():
    file = 'Sample_updated_type1.xlsx'
    wb_new_myra = load_workbook(filename=file)
    wb_orig_new_myra = xlrd.open_workbook(file)
    print(wb_new_myra.sheetnames)
    for num_sheet in wb_new_myra.sheetnames:
        if num_sheet == 'Base':
            ws_new_myra = wb_new_myra[num_sheet]
            sheet_new_myra = wb_orig_new_myra.sheet_by_name(num_sheet)
        elif num_sheet == 'To be Mapp - set 1':
            ws_set1_sheet = wb_new_myra[num_sheet]
            sheet_set1_sheet = wb_orig_new_myra.sheet_by_name(num_sheet)

    if ws_new_myra:
        start_column_ws = 2
        start_column_ws2 = 3
        start_Varient = 4
        Fuel_Type = 8
        fuel_type2 = 10
        Set1_id = 9
        Set2_id =9
        Id = 1
        Per_c=10 # 12
        start_row_ws_i = 2
        for i in range(sheet_new_myra.nrows):
            temp_Make_Name = ws_new_myra.cell(row=start_row_ws_i, column=start_column_ws).value
            temp_Model_Name = ws_new_myra.cell(row=start_row_ws_i, column=start_column_ws2).value
            temp_Variant_Name = ws_new_myra.cell(row=start_row_ws_i, column=start_Varient).value
            temp_Fuel_Type = ws_new_myra.cell(row=start_row_ws_i, column=Fuel_Type).value
            temp_Set1_id = ws_new_myra.cell(row=start_row_ws_i, column=Set2_id).value
            start_row_ws = 2
            List_fi = []
            List_fi_VariR = []
            List_fi21_VarIR = []
            List_fi2 = []
            List_fi1 = []
            List_fi_prob = []
            List_fi_prob_Vari = []
            LessProb_VarIR = []
            List_fi21 = []
            DirectMatc = []
            LessProb = []
            List_fi2_VarIR = []
            for j in range(sheet_set1_sheet.nrows):
                print(start_row_ws_i, start_row_ws)
                temp_MAKE_NAME = ws_set1_sheet.cell(row=start_row_ws, column=start_column_ws).value
                temp_MODEL_NAME = ws_set1_sheet.cell(row=start_row_ws, column=start_column_ws2).value
                temp_VARIANT_NAME = ws_set1_sheet.cell(row=start_row_ws, column=start_Varient).value
                temp_Fuel_Type2 = ws_set1_sheet.cell(row=start_row_ws, column=Fuel_Type).value
                temp_Id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                print('####################################Start###########################################')
                print(temp_Make_Name)
                print(temp_MAKE_NAME)
                print(temp_Model_Name)
                print(temp_MODEL_NAME)
                print(temp_Variant_Name)
                print(temp_VARIANT_NAME)
                print(temp_Set1_id)
                print(temp_Id)
                print('########################################End##########################################')
                temp_varTrFa = False
                if (str(temp_Make_Name).replace('MAHINDRA SSANGYONG', 'MAHINDRA').replace('MARUTI SUZUKI',
                                                                                          'MARUTI').replace('AND',
                                                                                                            '&').replace(
                        'MAHINDRA & MAHINDRA', 'MAHINDRA').lower().strip() == str(temp_MAKE_NAME).replace('AND',
                                                                                                          '&').replace(
                        'MAHINDRA & MAHINDRA', 'MAHINDRA').replace('MARUTI SUZUKI', 'MARUTI').replace('  ',
                                                                                                      ' ').lower().strip()) and (
                str(temp_MODEL_NAME).upper().strip().__contains__(str(temp_Model_Name).upper().strip())):
                    if (str(temp_VARIANT_NAME).replace('.', '').replace('-', '').upper().replace('BHARAT',
                                                                                                 'BS').replace('   ',
                                                                                                               ' ').strip() == str(
                            temp_Variant_Name).replace('.', '').replace('-', '').upper().strip()) and (
                    str(temp_Fuel_Type2).replace('.', '').replace('-', '').lower().strip().__contains__(
                            str(temp_Fuel_Type).replace('.', '').replace('-', '').lower().strip())):
                        print('jshbndhdhhhehhed')
                        print('sdg')
                        print('@@@@@@@@@@@@@@@@@@@@@@@Start@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
                        print(Set2_id)
                        print(Id)
                        print('@@@@@@@@@@@@@@@@@@@@@@@@End@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
                        te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                        if te_id not in DirectMatc:
                            DirectMatc.append(te_id)
                        List_fi = []
                        List_fi2 = []
                        List_fiContains = []
                        break
                    elif (
                    str(temp_Variant_Name).replace('.', '').replace('-', '').upper().replace('BHARAT', 'BS').replace(
                            '   ', ' ').strip().__contains__(
                            str(temp_VARIANT_NAME).replace('.', '').replace('-', '').upper().strip())) and (
                    str(temp_Fuel_Type2).replace('.', '').replace('-', '').lower().strip().__contains__(
                            str(temp_Fuel_Type).replace('.', '').replace('-', '').lower().strip())):
                        print('s,sdhcjhagcjgsjhcdshdchjsjdc')
                        te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value

                        List_fi_prob.append(te_id)
                        List_fi_prob_Vari.append(temp_VARIANT_NAME)

                        # if te_id not in List_fi_prob:
                        #     List_fi_prob.append(te_id)
                        # if temp_VARIANT_NAME not in List_fi_prob_Vari:
                        #     List_fi_prob_Vari.append(temp_VARIANT_NAME)
                    elif (
                    str(temp_VARIANT_NAME).replace('.', '').replace('-', '').upper().replace('BHARAT', 'BS').replace(
                            '   ', ' ').strip().__contains__(
                            str(temp_Variant_Name).replace('.', '').replace('-', '').upper().strip())) and (
                    str(temp_Fuel_Type2).replace('.', '').replace('-', '').lower().strip().__contains__(
                            str(temp_Fuel_Type).replace('.', '').replace('-', '').lower().strip())):
                        print('ssdkhbnadjna')
                        te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                        List_fi_prob.append(te_id)
                        List_fi_prob_Vari.append(temp_VARIANT_NAME)
                        # if te_id not in List_fi_prob:
                        #     List_fi_prob.append(te_id)
                        # if temp_VARIANT_NAME not in List_fi_prob_Vari:
                        #     List_fi_prob_Vari.append(temp_VARIANT_NAME)
                    else:
                        print('ssdkhassdbnadjna')
                        temp_Var = str(temp_Variant_Name).replace('.', '').replace('-', '').replace('K10', '').replace(
                            '1.1', '').replace('  ', '').upper().strip().split(' ')
                        V_J = str(temp_VARIANT_NAME).replace('.', '').replace('-', '').upper().replace('DI',
                                                                                                       '').replace(' ',
                                                                                                                   '').replace(
                            'STANDARD', 'STD').strip()
                        temp_ftev = False
                        print('temp_Var', temp_Var)
                        for i in temp_Var:
                            if i.upper().strip().__contains__(V_J):
                                if (str(temp_Fuel_Type2).replace('.', '').replace('-', '').lower().strip().__contains__(
                                        str(temp_Fuel_Type).replace('.', '').replace('-', '').lower().strip())):
                                    temp_ftev = True
                                    print('%%%%%%%%%%%%%%%%%%%Start%%%%%%%%%%%%%%%%%%%')
                                    print(temp_Set1_id)
                                    print(temp_Id)
                                    print('%%%%%%%%%%%%%%%%%%%End%%%%%%%%%%%%%%%%%%%')
                                    te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                    List_fi.append(te_id)
                                    List_fi_VariR.append(temp_VARIANT_NAME)
                                    # if te_id not in List_fi:
                                    #     List_fi.append(te_id)
                                    # if temp_VARIANT_NAME not in List_fi_VariR:
                                    #     List_fi_VariR.append(temp_VARIANT_NAME)
                        if not temp_ftev:
                            Upperss = False
                            temp_Var = str(temp_Variant_Name).replace('.', '').replace('-', '').replace('K10',
                                                                                                        '').replace(
                                '1.1', '').replace('  ', '').upper().strip()
                            V_J = str(temp_VARIANT_NAME).replace('AT 1.2', '').replace('1.2', '').replace('1.1',
                                                                                                          '').replace(
                                '.', '').replace('-', '').upper().replace('DI', '').replace('STANDARD',
                                                                                            'STD').strip().split(' ')
                            print('sdvkkns')
                            print("V_J", V_J)
                            for i in V_J:
                                print('sdvkknssxf')
                                print(i)
                                if i.upper().strip().__contains__(temp_Var.upper().strip()):
                                    Upperss = True
                                    print('jhsdgdcjhdcjshjhd')
                                    te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                    List_fi21.append(te_id)
                                    List_fi21_VarIR.append(temp_VARIANT_NAME)
                                    # if te_id not in List_fi21:
                                    #     List_fi21.append(te_id)
                                    # if temp_VARIANT_NAME not in List_fi21_VarIR:
                                    #     List_fi21_VarIR.append(temp_VARIANT_NAME)
                        if not Upperss:
                            print('jsefhgjhgwejfb sdgfhvjsgf')
                            temp_Var = str(temp_Variant_Name).replace('.', '').replace('-', '').replace('K10',
                                                                                                        '').replace(
                                '1.1', '').replace('  ', '').upper().strip().split(' ')
                            V_J = str(temp_VARIANT_NAME).replace('AT 1.2', '').replace('1.2', '').replace('1.1',
                                                                                                          '').replace(
                                '.', '').replace('-', '').upper().replace('DI', '').replace('STANDARD', 'STD').strip()
                            print("LessProb", V_J)
                            for i in range(len(temp_Var)):
                                print(i)
                                if len(temp_Var) >= 2:
                                    if V_J.upper().strip().__contains__(
                                            temp_Var[0].upper().strip()) and V_J.upper().strip().__contains__(
                                            temp_Var[1].upper().strip()):
                                        te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                        LessProb.append(te_id)
                                        LessProb_VarIR.append(temp_VARIANT_NAME)

                                        # if te_id not in LessProb:
                                        #     LessProb.append(te_id)
                                        # if temp_VARIANT_NAME not in LessProb_VarIR:
                                        #     LessProb_VarIR.append(temp_VARIANT_NAME)
                                    else:
                                        if V_J.upper().strip().__contains__(temp_Var[0].upper().strip()):
                                            te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                            LessProb.append(te_id)
                                            LessProb_VarIR.append(temp_VARIANT_NAME)

                                            # if te_id not in LessProb:
                                            #     LessProb.append(te_id)
                                            # if temp_VARIANT_NAME not in LessProb_VarIR:
                                            #     LessProb_VarIR.append(temp_VARIANT_NAME)
                                elif V_J.upper().strip().__contains__(temp_Var[0].upper().strip()):
                                    te_id = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                                    LessProb.append(te_id)
                                    LessProb_VarIR.append(temp_VARIANT_NAME)
                                    # if te_id not in LessProb:
                                    #     LessProb.append(te_id)
                                    # if temp_VARIANT_NAME not in LessProb_VarIR:
                                    #     LessProb_VarIR.append(temp_VARIANT_NAME)
                        if not List_fi and not LessProb:
                            te_ids = ws_set1_sheet.cell(row=start_row_ws, column=Id).value
                            List_fi2.append(te_ids)
                            List_fi2_VarIR.append(temp_VARIANT_NAME)
                            # if te_ids not in List_fi2:
                            #     List_fi2.append(te_ids)
                            # if temp_VARIANT_NAME not in List_fi2_VarIR:
                            #     List_fi2_VarIR.append(temp_VARIANT_NAME)

                start_row_ws = start_row_ws + 1
            if DirectMatc:
                PerCent = '100'
                print("DirectMatc", DirectMatc)
                List_fi = []
                List_fi2 = []
                List_fi1 = []
                List_fi_prob = []
                va = ''
                for i in DirectMatc:
                    tem_t = False
                    if not va:
                        tem_t = True
                        va = str(i) + '(' + str(PerCent) + ')'
                    if va and not tem_t:
                        va = va + ', ' + str(i) + '(' + str(PerCent) + ')'
                print(va)
                ws_new_myra.cell(row=start_row_ws_i, column=Per_c).value = PerCent
                ws_new_myra.cell(row=start_row_ws_i, column=Set2_id).value = va
            elif List_fi:
                print('List_fi')
                List_fi2 = []
                List_fi1 = []
                List_fi_prob = []
                SET_IDs = List_fi
                Vari_list=List_fi_VariR
                Temp_var=temp_Variant_Name

                va,Out_P=get_fina_value(SET_IDs,Vari_list,Temp_var)
                ws_new_myra.cell(row=start_row_ws_i, column=Per_c).value = Out_P
                ws_new_myra.cell(row=start_row_ws_i, column=Set2_id).value = va
            elif List_fi21:
                print('List_fi21')
                List_fi2 = []
                List_fi1 = []
                List_fi_prob = []
                SET_IDs = List_fi21
                Vari_list = List_fi21_VarIR
                Temp_var = temp_Variant_Name
                va,Out_P = get_fina_value(SET_IDs, Vari_list, Temp_var)
                ws_new_myra.cell(row=start_row_ws_i, column=Per_c).value = Out_P
                ws_new_myra.cell(row=start_row_ws_i, column=Set2_id).value = va
            elif List_fi_prob and not List_fi:
                print("List_fi_prob")
                List_fi2 = []
                List_fi = []
                SET_IDs = List_fi_prob
                Vari_list = List_fi_prob_Vari
                Temp_var = temp_Variant_Name
                va,Out_P = get_fina_value(SET_IDs, Vari_list, Temp_var)
                ws_new_myra.cell(row=start_row_ws_i, column=Per_c).value = Out_P
                ws_new_myra.cell(row=start_row_ws_i, column=Set2_id).value = va
            elif LessProb:
                print("LessProb")
                List_fi2 = []
                List_fi1 = []
                List_fi_prob = []
                SET_IDs = LessProb
                Vari_list = LessProb_VarIR
                Temp_var = temp_Variant_Name
                va,Out_P = get_fina_value(SET_IDs, Vari_list, Temp_var)
                ws_new_myra.cell(row=start_row_ws_i, column=Per_c).value = Out_P
                ws_new_myra.cell(row=start_row_ws_i, column=Set2_id).value = va
            else:
                if List_fi2:
                    print("List_fi2")
                    SET_IDs = List_fi2
                    Vari_list = List_fi2_VarIR
                    Temp_var = temp_Variant_Name
                    va,Out_P = get_fina_value(SET_IDs, Vari_list, Temp_var)
                    ws_new_myra.cell(row=start_row_ws_i, column=Per_c).value = Out_P
                    ws_new_myra.cell(row=start_row_ws_i, column=Set2_id).value = va
            start_row_ws_i = start_row_ws_i + 1
        wb_new_myra.save(filename=file)
        wb_new_myra.close()

def get_Duplicate(List1):
    items = []
    for item, count in collections.Counter(List1).items():
        if count > 1:
            items.append(item)
    return items

def get_percentage(str_i,List1):
    temp=[]
    temp1 = ''
    l = len(List1)
    z = process.extract(str_i, List1, scorer=fuzz.ratio, limit=l)
    print("z", z)
    temp1 = ''
    for i in range(len(List1)):
        for j in range(len(z)):
            if List1[i] in z[j][0]:
                temp1 = z[j][1]
        if temp1!='':
            temp.append(temp1)
    print("list1", List1)
    print("temp", temp)
    return temp

def get_fina_value(SET_IDs,Vari_list,Temp_var):
    PerCentage_ = get_percentage(Temp_var, Vari_list)
    Temp_list = get_Duplicate(SET_IDs)
    print("SET_IDs", SET_IDs)
    print("Temp_var",Temp_var)
    print("Vari_list",Vari_list)
    print("PerCentage_",PerCentage_)
    print("Temp_list",Temp_list)
    l_var = []
    l_id = []
    l_Per = []
    for i in range(len(Temp_list)):
        temp_str_var = ''
        temp_str_id = ''
        temp_str_Per = ''
        temp_percent = 0
        for j in range(len(SET_IDs)):
            if str(Temp_list[i]) in str(SET_IDs[j]):
                if PerCentage_[j] >= temp_percent:
                    temp_str_var = Vari_list[j]
                    temp_str_id = SET_IDs[j]
                    temp_str_Per = PerCentage_[j]
                    temp_percent = PerCentage_[j]
        if temp_str_var !='' and temp_str_id!='' and temp_str_Per!='':
            l_var.append(temp_str_var)
            l_id.append(temp_str_id)
            l_Per.append(temp_str_Per)
    for i in range(len(Temp_list)):
        for j in range(len(SET_IDs)):
            if str(Temp_list[i]) in str(SET_IDs[j]):
                SET_IDs[j] = ''
                Vari_list[j] = ''
                PerCentage_[j] = ''
    for i in range(len(SET_IDs)):
        if str(SET_IDs[i]) != '':
            l_var.append(Vari_list[i])
            l_id.append(SET_IDs[i])
            l_Per.append(PerCentage_[i])
    va = ''
    if l_Per:
        in_Percentage=l_Per
        in_Percentage.sort()
        Out_P=in_Percentage[-1]
    for i in range(len(l_id)):
        tem_t = False
        if not va:
            tem_t = True
            va = str(l_id[i]) + '(' + str(l_Per[i]) + ')'
        if va and not tem_t:
            va = va + ', ' + str(l_id[i]) + '(' + str(l_Per[i]) + ')'
    return va,Out_P

get_final()

